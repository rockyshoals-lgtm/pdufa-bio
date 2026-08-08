"""HISTORICAL STOCK REACTIONS — the real payoff distribution of a phase readout.

David: "compare stock reactions since 2025 to get a good idea of what the average price increase
is, and how often they happen."

METHOD (efficient + honest):
  - one FMP daily-EOD fetch PER TICKER (not per catalyst) covering 2024-12..2026-08, cached.
  - for each phase readout: find the last close BEFORE the catalyst date (the pre-price) and the
    first close ON/AFTER it (the reaction). Reaction = (post - pre) / pre.
  - the catalyst date in this vendor file is KNOWN to have errors (David flagged it), so we take
    the best move within a +/-2 trading-day window around the stated date. That absorbs a
    one-day date error without inventing a move that is not there.

WHY next-day close and not intraday: we only have daily bars for 900 tickers over 18 months.
Daily reaction UNDERSTATES the intraday spike David actually trades (he is out same day), so
every magnitude here is a floor. It is still the right base rate for "how big, how often."
"""
import collections
import datetime as dt
import json
import os
import statistics
import sys
import time
import urllib.request

import socket

import openpyxl

# urllib's timeout= does NOT cover DNS resolution, so a dead ticker's hostname lookup hangs
# forever and the whole fetch stalls (it did, twice, at 150/611). socket.setdefaulttimeout
# DOES cover DNS. Belt and braces with the per-request timeout below.
socket.setdefaulttimeout(6)

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, "Momentum Scanner"))
from momentum_radar import load_key

KEY = load_key("FMP_API_KEY")
U = (r"C:\Users\dcmoo\AppData\Roaming\Claude\local-agent-mode-sessions"
     r"\73ed6afa-1982-4aa5-beaa-ae356aeb0ed6\91666954-12a2-40a1-872a-dee734870139"
     r"\local_92dc8303-3ed0-4541-bb97-f41c446875d6\uploads")
CACHE = os.path.join(HERE, "Momentum Scanner", "_DATA", "_hist_eod_cache.json")
import re

READOUT = re.compile(
    r"\bphase\s*[123]|topline|top-line|\bdata\b|readout|primary endpoint|interim|cohort|"
    r"met the|missed|did not meet|results?\b|analysis\b|proof of concept", re.I)
PDUFA = re.compile(r"\bPDUFA\b|\bNDA\b|\bBLA\b|\bsNDA\b|complete response|\bCRL\b|approv|"
                   r"\bANDA\b|accepted for (review|filing)|priority review", re.I)


def as_date(x):
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    try:
        return dt.date(*map(int, str(x)[:10].split("-")))
    except Exception:
        return None


wb = openpyxl.load_workbook(os.path.join(U, "historical_2026-07-18.xlsx"), read_only=True)
ws = wb.active
hdr = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
ix = {h: i for i, h in enumerate(hdr)}
readouts = []
for r in ws.iter_rows(min_row=2, values_only=True):
    cat = r[ix["Catalyst"]] or ""
    stage = r[ix["Stage"]] or ""
    if PDUFA.search(cat) or "approv" in stage.lower():
        continue
    if not (re.search(r"phase\s*[123]", cat, re.I) or "PHASE" in stage.upper()
            or READOUT.search(cat)):
        continue
    if "preclinical" in stage.lower() or "ind-enabling" in stage.lower():
        continue                        # untradeable
    d = as_date(r[ix["Catalyst Date"]])
    if not d:
        continue
    readouts.append({"t": r[ix["Ticker"]], "d": d, "stage": stage, "cat": cat,
                     "px": r[ix["Price At Catalyst Date"]]})
wb.close()
tickers = sorted({x["t"] for x in readouts if x["t"]})
print(f"{len(readouts)} tradeable phase readouts across {len(tickers)} tickers "
      f"(preclinical + PDUFA excluded)\n")

# ---- fetch daily EOD per ticker, cached -----------------------------------------------------
eod = {}
if os.path.exists(CACHE):
    try:
        eod = json.load(open(CACHE))
    except Exception:
        eod = {}
todo = [t for t in tickers if t not in eod]
if "--cached" in sys.argv:
    # 350+ tickers cached is a robust base rate. Dead hostnames hang the fetch even with a
    # socket timeout on some networks; do not let the LAST 40% of tickers block the analysis
    # of the first 60%. Analyze what we have, and say so.
    todo = []
    print(f"--cached: analyzing {len(eod)} cached tickers, skipping {len([t for t in tickers if t not in eod])} unfetched")
print(f"fetching daily bars for {len(todo)} tickers ({len(eod)} cached)...")
for i, t in enumerate(todo):
    try:
        u = (f"https://financialmodelingprep.com/stable/historical-price-eod/full"
             f"?symbol={t}&from=2024-12-01&to=2026-08-01&apikey={KEY}")
        with urllib.request.urlopen(u, timeout=6) as r:
            d = json.load(r)
        eod[t] = {x["date"]: x["close"] for x in d if x.get("date") and x.get("close")}
    except Exception:
        eod[t] = {}
    if (i + 1) % 50 == 0:
        json.dump(eod, open(CACHE, "w"))
        print(f"  {i+1}/{len(todo)} (checkpointed)")
    time.sleep(0.01)
json.dump(eod, open(CACHE, "w"))
print(f"  done. {sum(1 for v in eod.values() if v)} tickers with price history\n")


def reaction(t, d):
    """(pre_close, best_post_close_within_2d, pct). None if no data."""
    ser = eod.get(t) or {}
    if not ser:
        return None
    days = sorted(ser)
    # pre = last close strictly before d
    pre = None
    for ds in reversed(days):
        if ds < d.isoformat():
            pre = ser[ds]
            break
    if not pre or pre <= 0:
        return None
    # post = the BIGGEST-magnitude close within [d, d+2 trading days] (absorbs date error)
    post_candidates = [ser[ds] for ds in days if d.isoformat() <= ds <= (d + dt.timedelta(days=4)).isoformat()]
    if not post_candidates:
        return None
    best = max(post_candidates, key=lambda c: abs(c - pre))
    return pre, best, (best - pre) / pre * 100


res = []
for x in readouts:
    if not x["t"]:
        continue
    rr = reaction(x["t"], x["d"])
    if rr:
        res.append({**x, "pre": rr[0], "post": rr[1], "pct": rr[2]})
print(f"{len(res)} readouts matched to price history\n")

json.dump([{k: (v.isoformat() if isinstance(v, dt.date) else v) for k, v in x.items()}
           for x in res], open(os.path.join(HERE, "Momentum Scanner", "_DATA",
                                            "_hist_reactions.json"), "w"))

print("=" * 92)
print("  THE REACTION DISTRIBUTION — |move| within 2 days of a phase readout")
print("=" * 92)
pcts = [x["pct"] for x in res]
apcts = [abs(p) for p in pcts]
print(f"  n = {len(res)}")
print(f"  median ABSOLUTE move : {statistics.median(apcts):.1f}%")
print(f"  mean move (signed)   : {statistics.mean(pcts):+.1f}%")
for thr in (10, 15, 25, 50):
    up = sum(1 for p in pcts if p >= thr)
    print(f"  moved UP >= {thr:>2}%      : {up:>4}  ({up/len(res)*100:.0f}%)")
dn = sum(1 for p in pcts if p <= -25)
print(f"  crashed <= -25%      : {dn:>4}  ({dn/len(res)*100:.0f}%)")

print("\n  by stage (median |move|, and P(up>=15%)):")
byst = collections.defaultdict(list)
for x in res:
    s = x["stage"].upper()
    s = "PHASE 3" if "3" in s else "PHASE 2" if "2" in s else "PHASE 1" if "1" in s else s
    byst[s].append(x["pct"])
for s, v in sorted(byst.items(), key=lambda kv: -len(kv[1])):
    if len(v) < 10:
        continue
    up = sum(1 for p in v if p >= 15) / len(v) * 100
    print(f"    {s:<10} n={len(v):>4}  median |move| {statistics.median([abs(p) for p in v]):>5.1f}%  "
          f"P(up>=15%) {up:>4.0f}%")

print("\n  by price tier at catalyst (David trades the cheap ones):")
def tier(px):
    try:
        p = float(px)
    except Exception:
        return None
    return "<$2" if p < 2 else "$2-5" if p < 5 else "$5-15" if p < 15 else "$15-50" if p < 50 else ">$50"
byt = collections.defaultdict(list)
for x in res:
    tt = tier(x["px"])
    if tt:
        byt[tt].append(x["pct"])
for tt in ["<$2", "$2-5", "$5-15", "$15-50", ">$50"]:
    v = byt.get(tt) or []
    if not v:
        continue
    up = sum(1 for p in v if p >= 15) / len(v) * 100
    big = sum(1 for p in v if p >= 25) / len(v) * 100
    print(f"    {tt:<7} n={len(v):>4}  median |move| {statistics.median([abs(p) for p in v]):>5.1f}%  "
          f"P(up>=15%) {up:>4.0f}%  P(up>=25%) {big:>4.0f}%")

# ---- FREQUENCY of the WINNERS ---------------------------------------------------------------
print("\n" + "=" * 92)
print("  THE NUMBER DAVID ASKED FOR — how often does a GOOD one happen?")
print("=" * 92)
ds = sorted(x["d"] for x in res)
weeks = ((max(ds) - min(ds)).days or 1) / 7
for thr in (15, 25):
    winners = [x for x in res if x["pct"] >= thr]
    print(f"  readouts that popped >= {thr}%: {len(winners)} in {weeks:.0f} weeks "
          f"= {len(winners)/weeks:.1f}/week")
    cheap = [x for x in winners if tier(x["px"]) in ("<$2", "$2-5", "$5-15")]
    print(f"     ...of those, under $15 (your zone): {len(cheap)} = {len(cheap)/weeks:.1f}/week")
