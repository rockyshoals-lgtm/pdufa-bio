"""BPC 8/22 export vs OUR 8/22 run — who has the sharper date, and where do we disagree?

Three questions, in the order they matter for preload + publishing on pdufa.bio:
  1. Does BPC carry REAL day precision, or the same quarter-bucket placeholders we just caught?
  2. Which forward names does BPC have that we MISS entirely (the coverage gap)?
  3. Where we BOTH have a name, do the dates AGREE? A disagreement is either their slip or ours.
"""
import csv, os, re, sys, collections, datetime
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.abspath(__file__))
TODAY = datetime.date(2026, 8, 22)

wb = load_workbook(os.path.join(HERE, "bpc_data", "fda_2026-08-22.xlsx"), read_only=True)
ws = wb.worksheets[0]
rows, hdr = [], None
for r in ws.iter_rows(values_only=True):
    if hdr is None:
        hdr = [str(x or "").strip() for x in r]
        continue
    rows.append(dict(zip(hdr, r)))
print(f"BPC sheet '{ws.title}': {len(rows)} rows")
print("columns:", hdr)


def cell(d, *names):
    for n in names:
        if n in d and d[n] not in (None, ""):
            return d[n]
    return None


# normalise
B = []
for d in rows:
    tk = str(cell(d, "Ticker", "ticker", "Symbol") or "").strip().upper()
    dt_ = cell(d, "Catalyst Date", "Date", "catalyst_date")
    stage = str(cell(d, "Stage", "Phase", "stage") or "")
    if not tk:
        continue
    ds = ""
    if isinstance(dt_, datetime.datetime):
        ds = dt_.date().isoformat()
    elif isinstance(dt_, datetime.date):
        ds = dt_.isoformat()
    elif dt_:
        ds = str(dt_)[:10]
    B.append({"tk": tk, "date": ds, "stage": stage,
              "drug": str(cell(d, "Drug", "drug", "Name") or "")[:32],
              "note": str(cell(d, "Catalyst", "Description", "Notes") or "")[:60]})
fwd = [b for b in B if b["date"] >= TODAY.isoformat()]
print(f"  {len(fwd)} forward-dated rows (>= {TODAY})\n")

# 1. PRECISION — does BPC clump on placeholder days like we do?
print("=" * 92)
print("  1. IS BPC's DATE A REAL DAY? (a real calendar does not clump)")
print("=" * 92)
cnt = collections.Counter(b["date"] for b in fwd)
for d, n in cnt.most_common(10):
    try:
        dd = datetime.date.fromisoformat(d)
        last = (datetime.date(dd.year + (dd.month == 12), dd.month % 12 + 1, 1)
                - datetime.timedelta(days=1)).day
        tag = ("MONTH-END bucket" if dd.day == last else
               "MID-MONTH bucket" if dd.day == 15 else "looks real")
    except Exception:
        tag = "?"
    print(f"   {d}  x{n:<4} {tag}")


def is_ph(d):
    try:
        dd = datetime.date.fromisoformat(d)
    except Exception:
        return True
    last = (datetime.date(dd.year + (dd.month == 12), dd.month % 12 + 1, 1)
            - datetime.timedelta(days=1)).day
    return dd.day in (1, 15, last)


ph = [b for b in fwd if is_ph(b["date"])]
real = [b for b in fwd if not is_ph(b["date"])]
print(f"\n   BPC forward rows: {len(real)} real-looking days, {len(ph)} placeholder-shaped "
      f"({100*len(ph)/max(1,len(fwd)):.0f}%)")

# 2/3. compare to ours
def load(p):
    fp = os.path.join(HERE, p)
    return list(csv.DictReader(open(fp, encoding="utf-8-sig", errors="replace"))) \
        if os.path.exists(fp) else []


F, C, G = load("readout_forward.csv"), load("readout_calendar.csv"), load("ctgov_readouts.csv")
ours = {}
for r in F:
    for col in ("window", "window_alt"):
        v = (r.get(col) or "").strip()
        if v:
            ours.setdefault(r["ticker"], set()).add(v)
for r in C:
    if (r.get("best_date") or "").strip():
        ours.setdefault(r["ticker"], set()).add(r["best_date"].strip())
for r in G:
    if (r.get("pcd") or "").strip():
        ours.setdefault(r.get("ticker"), set()).add(r["pcd"].strip())
ours.pop(None, None)
print(f"\n   our universe: {len(ours)} tickers with any date")

bt = {b["tk"] for b in fwd}
print("\n" + "=" * 92)
print("  2. COVERAGE GAP")
print("=" * 92)
missing = sorted(bt - set(ours))
print(f"   BPC forward names we have NO date for at all: {len(missing)}/{len(bt)}")
print("   ", ", ".join(missing[:45]))
extra = sorted(set(ours) - bt)
print(f"\n   names WE have that BPC's forward sheet does not: {len(extra)}")
print("   ", ", ".join(extra[:30]))

print("\n" + "=" * 92)
print("  3. WHERE WE BOTH HAVE A DATE — do we agree? (BPC real-day rows only)")
print("=" * 92)
agree = dis = 0
show = []
for b in sorted(real, key=lambda x: x["date"]):
    o = ours.get(b["tk"])
    if not o:
        continue
    if b["date"] in o:
        agree += 1
    else:
        dis += 1
        if b["date"] <= "2026-12-31":
            show.append((b["date"], b["tk"], b["stage"][:14], sorted(o)[:3]))
print(f"   exact-match {agree} | differ {dis}")
for d, tk, st, o in show[:25]:
    print(f"   BPC {d}  {tk:<7}{st:<15} ours: {', '.join(o)}")

# what BPC gives us that is a REAL day and we only have a bucket for
print("\n" + "=" * 92)
print("  4. THE PRIZE — BPC has a REAL DAY where we only have a bucket")
print("=" * 92)
BUCKET = re.compile(r"^(Q[1-4]|[12]H|MID)\s+20\d{2}$", re.I)
win = []
for b in sorted(real, key=lambda x: x["date"]):
    o = ours.get(b["tk"]) or set()
    if o and all(BUCKET.match(x) or is_ph(x) for x in o):
        win.append((b["date"], b["tk"], b["stage"][:16], b["drug"], sorted(o)[:2]))
print(f"   {len(win)} tickers upgradeable from bucket -> real day:")
for d, tk, st, dr, o in win[:30]:
    print(f"   {d}  {tk:<7}{st:<18}{dr:<26} was: {', '.join(o)}")
