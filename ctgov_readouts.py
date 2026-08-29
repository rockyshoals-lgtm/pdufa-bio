"""ctgov_readouts.py — the SPECIFIC-DATE readout pass, from ClinicalTrials.gov.

Why this exists (the SLS miss, 2026-07-18):
    readout_scan.py finds readouts by EDGAR full-text search for FORWARD phrases in the last
    45 days. That STRUCTURALLY misses a company like SLS (SELLAS) that guided its readout MONTHS
    ago and went quiet, or that never used our exact phrases. SLS was absent from every EDGAR
    pass — yet ClinicalTrials.gov has its lead trial (SLS009, NCT04588922) with a PRIMARY
    COMPLETION DATE of 2026-12-30. That date is the real "readout soon" signal, and it is far
    MORE SPECIFIC than "2H 2026".

What this does:
    For a watchlist of biotech tickers -> resolve to the company name (EDGAR ticker map) ->
    query CT.gov v2 by sponsor -> keep trials that are RECRUITING / ACTIVE_NOT_RECRUITING /
    COMPLETED-pending with a primary completion date in a forward window -> write a specific,
    dated readout candidate. This complements (does not replace) the EDGAR scan:
        EDGAR  = "the company SAID a readout is coming" (their words, vague dates)
        CT.gov = "the trial's data LOCKS on <date>"     (the calendar, specific dates)
    Run both; merge; dedupe by ticker.

Input tickers (union):
    - every ticker already in readout_forward.csv (adds specific CT.gov dates to EDGAR finds)
    - readout_watchlist.txt in this folder (one ticker per line) — seed it with names you KNOW
      have a readout coming (SLS is seeded). This is how you never miss a name you care about.

Honest limits: CT.gov primary completion dates are ESTIMATES and slip constantly (the [EST]
flag is shown). "Data locks Dec 30" is not "the stock moves Dec 30" — the PR can come weeks
after the lock, or the date can move. Verify against IR. Not investment advice.
"""
import collections
import csv
import datetime as dt
import json
import os
import sys
import time
import urllib.parse
import urllib.request
import re

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.abspath(__file__))
AGENT = os.environ.get("SEC_USER_AGENT", "David Moody rockyshoals@gmail.com")
FWD_MONTHS = 12          # keep PCDs from ~now out to +12 months
OVERDUE_DAYS = 180       # ...and recently-overdue ones, back to -180 days. Was 120: AMLX's LUCIDITY (PCD est. 2026-03) reported positive 2026-08-18, ~140d after its estimated PCD -- the 120-day cutoff had silently dropped the exact trial we needed. Deep-overdue rows sort to the bottom as STALE, so keeping them is cheap.
TICKER_MAP_CACHE = os.path.join(HERE, "bpc_data", "_edgar_ticker_map.json")


def _get(url, tries=3):
    for i in range(tries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": AGENT})
            with urllib.request.urlopen(req, timeout=20) as r:
                return r.read().decode("utf-8", "replace")
        except Exception:
            time.sleep(0.4 * (i + 1))
    return None


def ticker_to_company():
    """{TICKER: company title} from EDGAR's public map. Cached ~weekly."""
    try:
        if os.path.exists(TICKER_MAP_CACHE) and \
                (time.time() - os.path.getmtime(TICKER_MAP_CACHE)) < 7 * 86400:
            return json.load(open(TICKER_MAP_CACHE))
    except Exception:
        pass
    raw = _get("https://www.sec.gov/files/company_tickers.json")
    m = {}
    if raw:
        try:
            for v in json.loads(raw).values():
                t = (v.get("ticker") or "").upper()
                if t:
                    m[t] = v.get("title", "")
        except Exception:
            pass
        try:
            os.makedirs(os.path.dirname(TICKER_MAP_CACHE), exist_ok=True)
            json.dump(m, open(TICKER_MAP_CACHE, "w"))
        except Exception:
            pass
    return m


def _sponsor_query(name):
    """Clean a company title into a CT.gov sponsor search term.

    2026-07-18 FIX — was 'Tarsus Pharmaceuticals, Inc.' -> 'Tarsus', which matched TARSUS
    UNIVERSITY (a Turkish school) and returned its trials. Over-stripping to the first word
    destroys the disambiguating context. Now: drop ONLY the trailing legal-entity suffix and the
    comma, KEEP the descriptive words. 'Tarsus Pharmaceuticals, Inc.' -> 'Tarsus Pharmaceuticals'
    (matches the drug company, not the university). The lead-sponsor verification below is the
    safety net when a query still returns extras."""
    n = (name or "").split(",")[0].strip()
    for suf in (" Inc.", " Inc", " Corporation", " Corp.", " Corp", " Ltd.", " Ltd", " plc",
                " N.V.", " S.A.", " AG", " SE", " LLC", " Limited", " Co."):
        if n.endswith(suf):
            n = n[: -len(suf)].strip()
    return n or (name or "")


_ACADEMIC = ("university", "hospital", "institute", "college", "medical center", "clinic",
             "foundation", "national ", "ministry", "health system", "cancer center",
             "school of")


def _sponsor_ok(lead_name, company):
    """Does the trial's LEAD SPONSOR actually match the company (and is it not an academic
    center)? The safety net for a query that returns extras. 'Tarsus University' fails because
    it is academic; a real 'Tarsus Pharmaceuticals, Inc.' lead passes."""
    ld = (lead_name or "").lower()
    if not ld:
        return False
    comp_l = (company or "").lower()
    company_is_academic = any(a in comp_l for a in _ACADEMIC)
    if not company_is_academic and any(a in ld for a in _ACADEMIC):
        return False                       # drug ticker but academic trial sponsor -> reject
    # share the company's distinctive first token (skip generic leading words)
    toks = [w for w in re.findall(r"[a-z]+", comp_l) if len(w) > 2
            and w not in ("the", "inc", "corp", "ltd", "group", "holdings")]
    key = toks[0] if toks else ""
    return bool(key) and key in ld


def ctgov_studies(sponsor):
    q = urllib.parse.quote(sponsor)
    url = (f"https://clinicaltrials.gov/api/v2/studies?query.spons={q}"
           f"&fields=NCTId,BriefTitle,Phase,OverallStatus,PrimaryCompletionDate,"
           f"PrimaryCompletionDateType,LeadSponsorName&pageSize=40")
    raw = _get(url)
    if not raw:
        return []
    try:
        return json.loads(raw).get("studies", [])
    except Exception:
        return []


def _pcd_precision(raw):
    """DAY | MONTH | YEAR — how specific is the primary completion date REALLY?

    Restored 2026-08-29. Two separate ways a bucket ends up looking like a day:

      1. CT.gov itself only carries "2026-12" for most trials, and parse_pcd() below
         expands that to 2026-12-31. WE manufacture that day. It is not in the registry.
      2. Even full YYYY-MM-DD entries cluster on the 1st, the 15th and month ends --
         the sponsor typed a bucket into a date field. 190 of 219 PCDs in the
         2026-08-28 run (87%) had one of those three shapes.

    So precision is read off the RAW string, before parse_pcd touches it, and a
    full date is still demoted when it lands on a placeholder day.
    """
    s = (raw or "").strip()
    if not s:
        return ""
    p = s.split("-")
    if len(p) == 1:
        return "YEAR"
    if len(p) == 2:
        return "MONTH"
    try:
        d = dt.date(int(p[0]), int(p[1]), int(p[2]))
    except Exception:
        return ""
    last = (dt.date(d.year + (d.month == 12), d.month % 12 + 1, 1) - dt.timedelta(days=1)).day
    return "MONTH" if d.day in (1, 15, last) else "DAY"


def parse_pcd(s):
    """'2026-12-30' or '2026-12' -> date (end-of-month if no day).

    NB: the month-only expansion here is why _pcd_precision() must read the RAW string.
    """
    if not s:
        return None
    p = s.split("-")
    try:
        if len(p) == 3:
            return dt.date(int(p[0]), int(p[1]), int(p[2]))
        if len(p) == 2:
            y, m = int(p[0]), int(p[1])
            last = [31, 29 if y % 4 == 0 else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1]
            return dt.date(y, m, last)
        return dt.date(int(p[0]), 12, 31)
    except Exception:
        return None


def watchlist_tickers():
    tks = set()
    # 1. everything already in readout_forward.csv -- FRESHEST of {base, _new}: the scan falls
    #    back to a _new sibling when Excel holds the base file open, and reading the stale base
    #    here would quietly feed yesterday's tickers into today's CT.gov pass.
    cands = [c for c in (os.path.join(HERE, "readout_forward.csv"),
                         os.path.join(HERE, "readout_forward_new.csv")) if os.path.exists(c)]
    if cands:
        p = max(cands, key=os.path.getmtime)
        for r in csv.DictReader(open(p, encoding="utf-8-sig")):
            if r.get("ticker"):
                tks.add(r["ticker"].upper())
    # 2. the explicit watchlist (seed with names you KNOW have a readout coming)
    wl = os.path.join(HERE, "readout_watchlist.txt")
    if os.path.exists(wl):
        for ln in open(wl, encoding="utf-8"):
            t = ln.strip().upper()
            if t and not t.startswith("#"):
                tks.add(t)
    # 3. BPC's forward-dated phase names (2026-08-13): the 2026-08-13 comparison showed BPC
    #    listing 298 forward phase-readout names of which the EDGAR pass covered 112 -- the
    #    45-day filing window STRUCTURALLY misses anyone who guided earlier (the SLS gap, at
    #    scale). The freshest fda_*.xlsx in bpc_data closes it: every forward-dated Phase row's
    #    ticker joins the CT.gov pass, so the trials get OUR dating and confidence tags rather
    #    than trusting BPC's quarter-bucket dates.
    try:
        import glob as _g
        from openpyxl import load_workbook
        xs = sorted(_g.glob(os.path.join(HERE, "bpc_data", "fda_*.xlsx")))
        if xs:
            ws = load_workbook(xs[-1], read_only=True).worksheets[0]
            hdr = None
            today = dt.date.today().isoformat()
            n0 = len(tks)
            for row in ws.iter_rows(values_only=True):
                if hdr is None:
                    hdr = [str(x or "") for x in row]
                    continue
                d = dict(zip(hdr, row))
                stage = str(d.get("Stage") or "")
                cd = str(d.get("Catalyst Date") or "")[:10]
                tkr = str(d.get("Ticker") or "").strip().upper()
                if stage.lower().startswith("phase") and cd >= today and tkr.isalpha():
                    tks.add(tkr)
            print(f"  watchlist: +{len(tks) - n0} forward phase names from "
                  f"{os.path.basename(xs[-1])}")
    except Exception as e:
        print(f"  [warn] BPC watchlist merge skipped: {e}")
    return sorted(tks)


def main():
    today = dt.date.today()
    lo = today - dt.timedelta(days=OVERDUE_DAYS)
    hi = today + dt.timedelta(days=FWD_MONTHS * 31)
    tks = watchlist_tickers()
    tmap = ticker_to_company()
    print("=" * 90)
    print(f"  CT.gov READOUT PASS — {len(tks)} watchlist tickers, PCD window "
          f"{lo} .. {hi}")
    print("=" * 90)
    out = []
    for i, tk in enumerate(tks):
        company = tmap.get(tk)
        if not company:
            continue
        best = None
        for s in ctgov_studies(_sponsor_query(company)):
            ps = s.get("protocolSection", {})
            stat = ps.get("statusModule", {})
            status = stat.get("overallStatus", "")
            if status in ("WITHDRAWN", "TERMINATED", "SUSPENDED"):
                continue
            # 2026-08-13 (owner spec: not enrollment, not closed -- company-guided readouts):
            #   COMPLETED        = the trial CLOSED; its data locked whenever it locked, and the
            #                      PCD tells you nothing about when the company will TALK. 7 of
            #                      115 rows in the fresh run were closed trials wearing dates.
            #   NOT_YET_RECRUITING = no patients yet; its PCD is pure fiction.
            # RECRUITING / ENROLLING_BY_INVITATION survive but are marked LOW confidence below:
            # their PCD moves with enrollment, which is exactly the artifact to distrust.
            if status in ("COMPLETED", "NOT_YET_RECRUITING"):
                continue
            # VERIFY the lead sponsor is really this company and not an academic center that
            # happens to share a word (the Tarsus University false-match).
            lead = (ps.get("sponsorCollaboratorsModule", {}) or {}).get("leadSponsor", {}) \
                .get("name", "")
            if not _sponsor_ok(lead, company):
                continue
            _raw_pcd = (stat.get("primaryCompletionDateStruct", {}) or {}).get("date", "")
            pcd = parse_pcd(_raw_pcd)
            if not pcd or not (lo <= pcd <= hi):
                continue
            phases = ",".join(ps.get("designModule", {}).get("phases", []) or [])
            if "PHASE1" not in phases and "PHASE2" not in phases and "PHASE3" not in phases:
                continue
            rec = {
                "ticker": tk, "company": company,
                "pcd": pcd.isoformat(),
                "pcd_type": (stat.get("primaryCompletionDateStruct", {}) or {}).get("type", ""),
                "status": status,
                "phase": phases,
                "nct": ps.get("identificationModule", {}).get("nctId", ""),
                "title": (ps.get("identificationModule", {}).get("briefTitle", "") or "")[:70],
                "days_to_pcd": (pcd - today).days,
                # ACTIVE_NOT_RECRUITING = fully enrolled, PCD is a data-lock estimate: MED.
                # Still-enrolling statuses: the PCD moves with enrollment: LOW. Nothing from
                # CT.gov is HIGH -- only the company's own guidance is.
                "pcd_conf": ("MED" if status == "ACTIVE_NOT_RECRUITING" else "LOW"),
                # pcd_precision (restored 2026-08-29): CT.gov stores a bucket as a day.
                # 190 of last night's 219 PCDs (87%) sat on the 1st, the 15th or a month
                # end -- that is the registry's placeholder convention, not a data lock on
                # that date. Only the other 13% are genuine days. Downstream must render a
                # MONTH row as "September 2026", never as "September 30".
                "pcd_precision": _pcd_precision(_raw_pcd),
                "raw_pcd": _raw_pcd,
            }
            # keep the SOONEST forward-or-recently-overdue trial per ticker
            if best is None or abs(rec["days_to_pcd"]) < abs(best["days_to_pcd"]):
                best = rec
        if best:
            out.append(best)
        if (i + 1) % 25 == 0:
            print(f"  {i+1}/{len(tks)}  ({len(out)} with a dated trial)")
        time.sleep(0.06)

    out.sort(key=lambda r: r["pcd"])
    dst = os.path.join(HERE, "ctgov_readouts.csv")
    # LOCK-RESISTANT WRITE. This whole pass (73 CT.gov queries, ~90s) was thrown away by a
    # PermissionError when ctgov_readouts.csv was open in Excel / locked by OneDrive. Write to a
    # tmp, atomic-replace with retries, and if still locked fall back to a sibling file rather
    # than losing 90 seconds of work. Same pattern smart_money_enrich already uses.
    cols = ["ticker", "pcd", "raw_pcd", "pcd_precision", "pcd_type", "pcd_conf",
            "days_to_pcd", "phase", "status", "nct",
            "company", "title"]
    tmp = dst + ".tmp"
    with open(tmp, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(out)
    ok = False
    for attempt in range(5):
        try:
            os.replace(tmp, dst)
            ok = True
            break
        except PermissionError:
            time.sleep(1.3 * (attempt + 1))
    if not ok:
        dst = os.path.splitext(dst)[0] + "_new.csv"
        os.replace(tmp, dst)
        print(f"  [ctgov] ctgov_readouts.csv was LOCKED (close it) -> wrote "
              f"{os.path.basename(dst)} instead. Data is safe.")

    _pp = collections.Counter(r.get("pcd_precision") or "(none)" for r in out)
    print(f"  pcd precision: " + "  ".join(f"{k}={n}" for k, n in _pp.most_common())
          + f"   <- only DAY is a real calendar date; MONTH/YEAR are buckets")

    print("\n" + "=" * 90)
    print(f"  {len(out)} tickers with a specific-dated trial readout -> ctgov_readouts.csv")
    print("=" * 90)
    print(f"  {'tkr':<6}{'PCD':<12}{'in':>6}{'phase':<14}{'status':<20}{'nct':<13}")
    for r in out[:40]:
        est = "~" if r["pcd_type"].upper().startswith("EST") else " "
        print(f"  {r['ticker']:<6}{est}{r['pcd']:<11}{r['days_to_pcd']:>5}d "
              f"{r['phase'][:13]:<14}{r['status'][:19]:<20}{r['nct']:<13}")
    print("\n  ~ = estimated PCD (slips constantly). Data can print weeks after the lock.")
    print("  Verify against IR. Not investment advice.")
    sls = [r for r in out if r["ticker"] == "SLS"]
    if sls:
        print(f"\n  SLS now caught: {sls[0]['nct']} PCD {sls[0]['pcd']} ({sls[0]['phase']})")


if __name__ == "__main__":
    main()
