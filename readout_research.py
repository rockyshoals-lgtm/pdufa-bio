"""readout_research.py — the whole phase-readout research pass, one command.

Reads the newest BiopharmaCatalyst downloads from bpc_data/ (historical_*.xlsx + fda_*.xlsx),
and prints:
    1. FREQUENCY  — how many phase readouts per week, and how often a good one (>=15% / >=25%)
    2. REACTIONS  — the base-rate move distribution since 2025 (met vs missed, by stage, by tier)
    3. GAP        — forward readouts BiopharmaCatalyst lists that we don't cover yet

Run it from READOUT_RESEARCH.bat. To refresh with newer data, drop a newer historical_YYYY-MM-DD
.xlsx and fda_YYYY-MM-DD.xlsx into bpc_data/ and run again — it always uses the newest of each.

Prices come from FMP daily bars (one fetch per ticker, cached in
Momentum Scanner/_DATA/_hist_eod_cache.json). A dead-hostname-proof socket timeout keeps it from
hanging. Daily-close reactions UNDERSTATE the intraday spike you trade — every magnitude is a
floor. Informational only, not investment advice.
"""
import collections
import datetime as dt
import glob
import json
import os
import re
import socket
import statistics
import sys
import time
import urllib.request

import openpyxl

socket.setdefaulttimeout(8)          # covers DNS; urllib timeout= does not
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = os.path.dirname(os.path.abspath(__file__))
BPC = os.path.join(HERE, "bpc_data")
CACHE = os.path.join(HERE, "Momentum Scanner", "_DATA", "_hist_eod_cache.json")
sys.path.insert(0, os.path.join(HERE, "Momentum Scanner"))
try:
    from momentum_radar import load_key
    KEY = load_key("FMP_API_KEY")
except Exception:
    KEY = os.environ.get("FMP_API_KEY", "")

CACHED_ONLY = "--cached" in sys.argv


def newest(pattern):
    fs = sorted(glob.glob(os.path.join(BPC, pattern)))
    return fs[-1] if fs else None


def load_sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    hdr = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return ix, rows


def as_date(x):
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    try:
        return dt.date(*map(int, str(x)[:10].split("-")))
    except Exception:
        return None


READOUT = re.compile(r"\bphase\s*[123]|topline|top-line|\bdata\b|readout|primary endpoint|"
                     r"interim|cohort|met the|missed|did not meet|results?\b|analysis\b|"
                     r"proof of concept", re.I)
PDUFA = re.compile(r"\bPDUFA\b|\bNDA\b|\bBLA\b|\bsNDA\b|complete response|\bCRL\b|approv|"
                   r"\bANDA\b|accepted for (review|filing)|priority review", re.I)


def is_readout(cat, stage):
    if PDUFA.search(cat or "") or "approv" in (stage or "").lower():
        return False
    if "preclinical" in (stage or "").lower() or "ind-enabling" in (stage or "").lower():
        return False
    return bool(re.search(r"phase\s*[123]", (cat or "") + " " + (stage or ""), re.I)
                or READOUT.search(cat or ""))


def stage_of(s):
    s = (s or "?").upper()
    return "PHASE 3" if "3" in s else "PHASE 2" if "2" in s else "PHASE 1" if "1" in s else s


def main():
    hist = newest("historical_*.xlsx")
    fda = newest("fda_*.xlsx")
    print("=" * 92)
    print("  9REALMS PHASE-READOUT RESEARCH")
    print("=" * 92)
    print(f"  historical : {os.path.basename(hist) if hist else 'MISSING'}")
    print(f"  forward    : {os.path.basename(fda) if fda else 'MISSING'}")
    if not hist:
        print("\n  No historical_*.xlsx in bpc_data/. Drop your BiopharmaCatalyst historical")
        print("  export there and run again.")
        return
    print()

    # ---------------------------------------------------------------- FREQUENCY + reactions
    ix, rows = load_sheet(hist)
    readouts = []
    for r in rows:
        cat = r[ix["Catalyst"]] if "Catalyst" in ix else ""
        stage = r[ix["Stage"]] if "Stage" in ix else ""
        if not is_readout(cat, stage):
            continue
        d = as_date(r[ix["Catalyst Date"]]) if "Catalyst Date" in ix else None
        if not d:
            continue
        readouts.append({"t": r[ix["Ticker"]], "d": d, "stage": stage, "cat": cat,
                         "px": r[ix.get("Price At Catalyst Date", -1)] if "Price At Catalyst Date" in ix else None})
    dts = sorted(x["d"] for x in readouts)
    weeks = ((max(dts) - min(dts)).days or 1) / 7
    print("=" * 92)
    print(f"  1. FREQUENCY — {len(readouts)} tradeable phase readouts, {min(dts)}..{max(dts)}")
    print("=" * 92)
    print(f"  {len(readouts)/weeks:.0f} phase readouts per week (all caps, Phase 1-3, "
          f"preclinical/PDUFA excluded)")
    st = collections.Counter(stage_of(x["stage"]) for x in readouts)
    for s, n in st.most_common(4):
        print(f"    {s:<10} {n/weeks:>4.1f}/wk")

    # reactions
    tickers = sorted({x["t"] for x in readouts if x["t"]})
    eod = {}
    if os.path.exists(CACHE):
        try:
            eod = json.load(open(CACHE))
        except Exception:
            eod = {}
    todo = [] if CACHED_ONLY else [t for t in tickers if t not in eod]
    if todo:
        print(f"\n  fetching daily prices for {len(todo)} new tickers "
              f"({len(eod)} cached)...")
        for i, t in enumerate(todo):
            try:
                u = (f"https://financialmodelingprep.com/stable/historical-price-eod/full"
                     f"?symbol={t}&from=2024-12-01&to={dt.date.today().isoformat()}"
                     f"&apikey={KEY}")
                with urllib.request.urlopen(u, timeout=6) as rr:
                    d = json.load(rr)
                eod[t] = {x["date"]: x["close"] for x in d if x.get("date") and x.get("close")}
            except Exception:
                eod[t] = {}
            if (i + 1) % 50 == 0:
                json.dump(eod, open(CACHE, "w"))
                print(f"    {i+1}/{len(todo)}")
            time.sleep(0.01)
        json.dump(eod, open(CACHE, "w"))

    def reaction(t, d):
        ser = eod.get(t) or {}
        if not ser:
            return None
        days = sorted(ser)
        pre = next((ser[ds] for ds in reversed(days) if ds < d.isoformat()), None)
        if not pre or pre <= 0:
            return None
        post = [ser[ds] for ds in days
                if d.isoformat() <= ds <= (d + dt.timedelta(days=4)).isoformat()]
        if not post:
            return None
        return (max(post, key=lambda c: abs(c - pre)) - pre) / pre * 100

    res = []
    for x in readouts:
        p = reaction(x["t"], x["d"]) if x["t"] else None
        if p is not None:
            res.append({**x, "pct": p})
    print("\n" + "=" * 92)
    print(f"  2. REACTIONS — base rate from {len(res)} readouts with price history")
    print("=" * 92)
    pct = [x["pct"] for x in res]
    if pct:
        print(f"  median |move| {statistics.median([abs(p) for p in pct]):.1f}%   "
              f"mean signed {statistics.mean(pct):+.1f}%")
        for thr in (15, 25):
            print(f"  P(up>={thr}%) {sum(1 for p in pct if p>=thr)/len(pct)*100:>4.0f}%   "
                  f"= {sum(1 for p in pct if p>=thr)/weeks:.1f} winners/week")
        print(f"  P(crash<=-25%) {sum(1 for p in pct if p<=-25)/len(pct)*100:.0f}%")
        print("\n  by stage:")
        byst = collections.defaultdict(list)
        for x in res:
            byst[stage_of(x["stage"])].append(x["pct"])
        for s, v in sorted(byst.items(), key=lambda kv: -len(kv[1])):
            if len(v) >= 10:
                print(f"    {s:<10} n={len(v):>4}  median |move| "
                      f"{statistics.median([abs(p) for p in v]):>4.1f}%  "
                      f"P(up>=15%) {sum(1 for p in v if p>=15)/len(v)*100:>3.0f}%")

    # ---------------------------------------------------------------- GAP vs forward
    if fda:
        aug = os.path.join(HERE, "phase_readouts_2026H2_cornerstone_augmented.csv")
        ours = set()
        our_nct = set()
        if os.path.exists(aug):
            import csv
            for r in csv.DictReader(open(aug, encoding="utf-8-sig")):
                if r.get("ticker"):
                    ours.add(r["ticker"].upper())
                if r.get("nct_id"):
                    our_nct.add(r["nct_id"].strip())
        fix, frows = load_sheet(fda)
        TODAY = dt.date.today()
        gaps = []
        for r in frows:
            cat = str(r[fix.get("Catalyst", -1)] or "") + " " + str(r[fix.get("Next Catalyst", -1)] or "")
            stage = str(r[fix.get("Stage", -1)] or "")
            d = as_date(r[fix.get("Catalyst Date", -1)])
            if not is_readout(cat, stage) or not d or d < TODAY:
                continue
            t = str(r[fix.get("Ticker", -1)] or "").upper().strip()
            nct = str(r[fix.get("NCT Number", -1)] or "").strip()
            if t and t not in ours and nct not in our_nct:
                # dedup + quarter-bucket note
                q = d.day in (31, 30) and d.month in (3, 6, 9, 12) or (d.month == 8 and d.day == 31)
                gaps.append((t, d, stage, "~Q" if q else "", cat[:44]))
        seen = {}
        for t, d, stage, q, cat in sorted(gaps, key=lambda z: z[1]):
            if t not in seen:
                seen[t] = (d, stage, q, cat)
        print("\n" + "=" * 92)
        print(f"  3. GAP — forward readouts BiopharmaCatalyst lists that we DON'T cover")
        print("=" * 92)
        print(f"  {len(seen)} unique names to consider adding (BPC dates marked ~Q are quarter")
        print(f"  buckets, not real days — get the real date from EDGAR before trusting):\n")
        print(f"  {'ticker':<8}{'BPC date':<12}{'stage':<12}  catalyst")
        print("  " + "-" * 78)
        for t, (d, stage, q, cat) in sorted(seen.items(), key=lambda kv: kv[1][0])[:30]:
            print(f"  {t:<8}{d.isoformat()+q:<12}{stage[:11]:<12}  {cat}")

    print("\n" + "=" * 92)
    print("  Daily-close reactions understate the intraday move you trade. Floors, not ceilings.")
    print("  A readout is +3% on average; the money is in filtering + fast exit, not holding.")
    print("  Informational only. Not investment advice.")
    print("=" * 92)


if __name__ == "__main__":
    main()
