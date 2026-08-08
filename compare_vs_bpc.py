#!/usr/bin/env python3
"""
compare_vs_bpc.py  --  head-to-head: OUR crawl vs BPC on PDUFAs, with a DATE-ACCURACY audit.

Dates must be 100% right for catalyst investing. This tool matches our primary-sourced
catalysts against BPC's list by ticker + drug, and reports three things:

  1) OUR WINS      PDUFAs WE have that BPC does NOT (we're ahead).
  2) BPC GAPS      PDUFAs BPC has that WE are missing (independently source these, then we win).
  3) DATE CONFLICTS  same drug in both, but the DATES DISAGREE. THESE ARE THE CRITICAL ONES —
                     each must be red-teamed against the primary source (company PR / SEC 8-K /
                     FDA). Our date carries its own source_url; BPC's does not. Whoever's date
                     matches the primary filing wins; fix ours if we're wrong, flag BPC if they are.

Writes compare_vs_bpc.csv (every row, with our source_url for verification) + prints a summary.

Usage:
  python compare_vs_bpc.py [--ours catalysts_out/catalysts_public.csv] [--bpc fda_2026-07-09.xlsx]
"""
import os, re, csv, sys, argparse, datetime
try:
    import openpyxl
except Exception:
    sys.exit("needs openpyxl:  python -m pip install openpyxl")

PDUFA_OURS = ("pdufa", "regulatory", "bla", "nda", "snda", "sbla", "gdufa", "fda decision", "approval", "crl")
PDUFA_BPC  = ("regulatory decision", "pdufa", "fda decision", "approval", "crl")

def drug_root(s):
    s = (s or "").lower()
    s = re.sub(r"\(.*?\)", " ", s)                 # drop trial names in parens
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    toks = [t for t in s.split() if len(t) > 2 and t not in
            ("the", "and", "for", "plus", "with", "formerly")]
    return toks[0] if toks else ""

def norm_date(x):
    if x is None: return ""
    if isinstance(x, (datetime.date, datetime.datetime)): return x.strftime("%Y-%m-%d")
    s = str(x).strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m: return m.group(0)
    # "August 2027" / "Q3 2026" -> month/quarter precision (kept as-is for display)
    return s[:10]

def load_ours(p):
    out = []
    for r in csv.DictReader(open(p, encoding="utf-8")):
        ct = (r.get("catalyst_type") or "").lower()
        if not any(t in ct for t in PDUFA_OURS): continue
        d = norm_date(r.get("catalyst_date"))
        if len(d) < 7: continue
        out.append(dict(ticker=(r.get("ticker") or "").upper().strip(),
                        root=drug_root(r.get("drug")), date=d,
                        drug=(r.get("drug") or "").strip(),
                        prec=(r.get("date_precision") or ""),
                        src=(r.get("source_url") or r.get("source") or "")))
    return out

def load_bpc(p):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]; rows = list(ws.iter_rows(values_only=True))
    H = {h: i for i, h in enumerate(rows[0])}
    def col(*names):
        for n in names:
            if n in H: return H[n]
        return None
    ti, di, ni, dr = col("Ticker"), col("Catalyst Date"), col("Next Catalyst"), col("Drug")
    out = []
    for r in rows[1:]:
        ty = (r[ni] or "").lower() if ni is not None else ""
        if not any(t in ty for t in PDUFA_BPC): continue
        d = norm_date(r[di]) if di is not None else ""
        if len(d) < 7: continue
        out.append(dict(ticker=(str(r[ti]) if ti is not None else "").upper().strip(),
                        root=drug_root(r[dr] if dr is not None else ""), date=d,
                        drug=(str(r[dr]) if dr is not None else "").strip()))
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ours", default="catalysts_out/catalysts_public.csv")
    ap.add_argument("--bpc", default="fda_2026-07-09.xlsx")
    a = ap.parse_args()
    if not os.path.exists(a.ours): sys.exit(f"our crawl not found: {a.ours} (run the crawler first)")
    if not os.path.exists(a.bpc): sys.exit(f"BPC file not found: {a.bpc}")

    ours = load_ours(a.ours); bpc = load_bpc(a.bpc)
    ours_k = {(o["ticker"], o["root"]): o for o in ours if o["root"]}
    bpc_k  = {(b["ticker"], b["root"]): b for b in bpc if b["root"]}

    wins   = [ours_k[k] for k in ours_k if k not in bpc_k]
    gaps   = [bpc_k[k]  for k in bpc_k  if k not in ours_k]
    both   = [k for k in ours_k if k in bpc_k]
    conflicts = [(ours_k[k], bpc_k[k]) for k in both if ours_k[k]["date"][:10] != bpc_k[k]["date"][:10]]
    agree  = [k for k in both if ours_k[k]["date"][:10] == bpc_k[k]["date"][:10]]

    print("="*70)
    print(f"PDUFA HEAD-TO-HEAD  —  ours {len(ours)}  vs  BPC {len(bpc)}")
    print(f"  matched drugs: {len(both)}  |  dates AGREE: {len(agree)}  |  DATE CONFLICTS: {len(conflicts)}")
    print(f"  OUR WINS (we have, BPC doesn't): {len(wins)}")
    print(f"  BPC GAPS (BPC has, we miss):     {len(gaps)}")
    verdict = "AHEAD" if len(ours) > len(bpc) and len(gaps) == 0 else ("AHEAD on count" if len(ours) > len(bpc) else "BEHIND — close the gaps")
    print(f"  VERDICT: {verdict}")
    print("="*70)
    if conflicts:
        print("\n!! DATE CONFLICTS — verify each vs the primary source (our date has a source_url) !!")
        for o, b in sorted(conflicts, key=lambda x: x[0]["date"]):
            print(f"  {o['ticker']:6} OURS {o['date']} ({o['prec']})  vs  BPC {b['date']}   {o['drug'][:40]}")
            print(f"         our source: {o['src'][:90]}")
    if gaps:
        print(f"\nBPC GAPS to independently source ({len(gaps)}):")
        for g in sorted(gaps, key=lambda x: x["date"]):
            print(f"  {g['ticker']:6} {g['date']}  {g['drug'][:50]}")

    with open("compare_vs_bpc.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(["class","ticker","our_date","our_precision","bpc_date","drug","our_source"])
        for o in wins:        w.writerow(["OUR_WIN", o["ticker"], o["date"], o["prec"], "", o["drug"], o["src"]])
        for g in gaps:        w.writerow(["BPC_GAP", g["ticker"], "", "", g["date"], g["drug"], ""])
        for o, b in conflicts:w.writerow(["DATE_CONFLICT", o["ticker"], o["date"], o["prec"], b["date"], o["drug"], o["src"]])
    print("\nwrote compare_vs_bpc.csv (every win/gap/conflict, with our source_url for verification)")

if __name__ == "__main__":
    main()
