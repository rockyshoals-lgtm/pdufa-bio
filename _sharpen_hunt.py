"""Hunt for date-sharpening signal we ALREADY own but do not use.

THREE HYPOTHESES, each testable on files already on disk:

H1  CONFERENCE NAMES IN OUR OWN EDGAR TEXT. A company that says "data will be presented at
    ESMO" has effectively given a hard date -- the congress agenda is public. We store that
    sentence in `context` and throw the conference name away. If the vague rows mention
    conferences at a decent rate, joining to a conference calendar converts HALF/QUARTER
    buckets straight to DAY.

H2  THE SLIP MODEL. BPC's historical file has 4,176 REALIZED catalyst dates. If a bucket
    like "Q4 2026" is systematically resolved late/early, we can publish a bucket as a
    DISTRIBUTION ("median Nov 12, 80% between Oct 15 and Dec 20") instead of a flat quarter.
    Measure where in its quarter/half a readout actually lands.

H3  CONFERENCE SEASONALITY BY THERAPEUTIC AREA. If oncology readouts cluster on ESMO/ASCO
    weeks and cardiology on ESC/AHA, then a vague oncology row in Q4 is far more likely to
    be ESMO week than a random Q4 day.
"""
import csv, os, re, sys, json, glob, collections, datetime, statistics
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------- H1
CONFS = {
    "ESMO": "European Society for Medical Oncology", "ASCO": "American Society of Clinical Onc",
    "ASH": "American Society of Hematology", "AACR": "American Association for Cancer Res",
    "ESC": "European Society of Cardiology", "AHA": "American Heart Association",
    "ACC": "American College of Cardiology", "EASD": "European Assoc Study of Diabetes",
    "ADA": "American Diabetes Association", "AAN": "American Academy of Neurology",
    "AAD": "American Academy of Dermatology", "EADV": "European Acad Dermatology",
    "EULAR": "European Alliance Rheumatology", "ACR": "American College of Rheumatology",
    "ERS": "European Respiratory Society", "ATS": "American Thoracic Society",
    "WCLC": "World Conference on Lung Cancer", "SITC": "Society for Immunotherapy of Cancer",
    "EHA": "European Hematology Association", "ASTRO": "Am Soc Radiation Oncology",
    "AASLD": "Am Assoc Study of Liver Diseases", "DDW": "Digestive Disease Week",
    "ARVO": "Assoc Research Vision Ophthalmology", "AAO": "Am Academy of Ophthalmology",
    "EURETINA": "EURETINA", "ASRS": "Am Society of Retina Specialists",
    "CTAD": "Clinical Trials on Alzheimer's", "AAIC": "Alzheimer's Assoc Intl Conf",
    "ASGCT": "Am Soc Gene & Cell Therapy", "SLEEP": "SLEEP", "ASN": "Am Society of Nephrology",
    "IDWeek": "IDWeek", "CROI": "CROI", "MSVirtual": "MS", "ECTRIMS": "ECTRIMS",
    "SABCS": "San Antonio Breast Cancer", "ESMO Breast": "ESMO Breast",
}
RX = re.compile(r"\b(" + "|".join(sorted(CONFS, key=len, reverse=True)) + r")\b")

F = list(csv.DictReader(open(os.path.join(HERE, "readout_forward.csv"),
                             encoding="utf-8-sig", errors="replace")))
vague = [r for r in F
         if (r.get("window_precision") or "") not in ("DAY",)
         and (r.get("window_alt_precision") or "") not in ("DAY",)]
hits = []
for r in vague:
    blob = " ".join([(r.get("context") or ""), (r.get("phrases") or ""),
                     (r.get("just_reported") or "")])
    m = RX.findall(blob)
    if m:
        hits.append((r["ticker"], sorted(set(m)), (r.get("window") or r.get("window_alt")),
                     (r.get("armed_lane") or "")))
print("=" * 92)
print("  H1 — CONFERENCE NAMES SITTING IN OUR OWN TEXT, ON UNDATED ROWS")
print("=" * 92)
print(f"  vague rows: {len(vague)}   with a conference named: {len(hits)} "
      f"({100*len(hits)/max(1,len(vague)):.0f}%)")
for tk, cs, w, lane in hits[:20]:
    print(f"    {tk:<7}{','.join(cs):<18}{str(w)[:12]:<14}{lane}")

# how often does the FULL forward file mention a conference anywhere?
anyc = sum(1 for r in F if RX.search(" ".join([(r.get("context") or ""),
                                               (r.get("phrases") or "")])))
print(f"\n  across ALL {len(F)} forward rows, {anyc} mention a conference")

# ---------------------------------------------------------------- H2 + H3
print("\n" + "=" * 92)
print("  H2 — WHERE INSIDE ITS QUARTER DOES A READOUT ACTUALLY LAND? (BPC historical)")
print("=" * 92)
hp = sorted(glob.glob(os.path.join(HERE, "bpc_data", "historical_*.xlsx")))
rows = []
if hp:
    ws = load_workbook(hp[-1], read_only=True).worksheets[0]
    hdr = None
    for r in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(x or "").strip() for x in r]
            continue
        rows.append(dict(zip(hdr, r)))
print(f"  {os.path.basename(hp[-1]) if hp else '-'}: {len(rows)} realized catalysts")


def dt_(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    try:
        return datetime.date.fromisoformat(str(v)[:10])
    except Exception:
        return None


real = [(dt_(d.get("Catalyst Date")), str(d.get("Stage") or ""), str(d.get("Conference") or ""),
         str(d.get("Ticker") or ""), str(d.get("Indication") or ""))
        for d in rows]
real = [x for x in real if x[0]]
print(f"  usable dated rows: {len(real)}")

# day-of-quarter distribution: is the quarter uniformly used, or back-loaded?
buckets = collections.Counter()
for d, stage, conf, tk, ind in real:
    q = (d.month - 1) // 3
    qstart = datetime.date(d.year, q * 3 + 1, 1)
    qend = (datetime.date(d.year + (q == 3), (q * 3 + 4 - 1) % 12 + 1, 1)
            - datetime.timedelta(days=1)) if q < 3 else datetime.date(d.year, 12, 31)
    span = (qend - qstart).days or 1
    pos = (d - qstart).days / span
    buckets[min(9, int(pos * 10))] += 1
tot = sum(buckets.values())
print("\n  position within the calendar quarter (decile, 0=first days .. 9=last days):")
for i in range(10):
    n = buckets.get(i, 0)
    bar = "█" * max(0, round(40 * n / max(1, max(buckets.values()))))
    print(f"    d{i}  {n:>5}  {100*n/max(1,tot):>5.1f}%  {bar}")

# month-of-year clustering -> conference seasonality
print("\n  catalysts by MONTH (conference season shows up here):")
mon = collections.Counter(d.month for d, *_ in real)
for m in range(1, 13):
    n = mon.get(m, 0)
    bar = "█" * max(0, round(38 * n / max(1, max(mon.values()))))
    print(f"    {datetime.date(2026,m,1).strftime('%b')}  {n:>5}  {bar}")

# H3: how many historical catalysts carried a conference, and which
print("\n" + "=" * 92)
print("  H3 — HOW OFTEN IS A REALIZED CATALYST A CONFERENCE EVENT?")
print("=" * 92)
withc = [x for x in real if x[2].strip()]
print(f"  {len(withc)}/{len(real)} ({100*len(withc)/max(1,len(real)):.0f}%) carried a "
      f"Conference value")
cc = collections.Counter()
for d, stage, conf, tk, ind in withc:
    m = RX.search(conf)
    cc[m.group(1) if m else conf[:26]] += 1
for k, n in cc.most_common(14):
    print(f"    {k:<30}{n}")
