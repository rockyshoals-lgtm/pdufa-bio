"""What data do we already own that could SHARPEN a date, and are we using it?"""
import os, sys, glob, json, csv
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.abspath(__file__))

print("=" * 92)
print("  A) BPC HISTORICAL — realized dates for catalysts that were once GUIDANCE")
print("=" * 92)
for p in sorted(glob.glob(os.path.join(HERE, "bpc_data", "historical_*.xlsx"))):
    ws = load_workbook(p, read_only=True).worksheets[0]
    hdr = None
    n = 0
    for r in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(x or "").strip() for x in r]
            continue
        n += 1
    print(f"  {os.path.basename(p)}: {n} rows")
    print(f"    columns: {hdr}")

print("\n" + "=" * 92)
print("  B) CT.GOV — fields we pull vs fields available")
print("=" * 92)
G = list(csv.DictReader(open(os.path.join(HERE, "ctgov_readouts.csv"),
                             encoding="utf-8-sig", errors="replace")))
print(f"  we store: {list(G[0].keys()) if G else '-'}")
print("  CT.gov v2 ALSO exposes (unused): CompletionDate, StudyFirstPostDate,")
print("     LastUpdatePostDate, EnrollmentCount, EnrollmentType, WhyStopped,")
print("     ResultsFirstPostDate, StatusVerifiedDate")

print("\n" + "=" * 92)
print("  C) ODIN training set — 2,200+ PDUFA events with REALIZED dates (ground truth)")
print("=" * 92)
for name in ("ODIN_MODEL_READY_v1071_T1_2015on_ENRICHED.csv",
             "ODIN_MODEL_READY_v1071_ENRICHED_v2.csv"):
    p = os.path.join(HERE, name)
    if os.path.exists(p):
        with open(p, encoding="utf-8", errors="replace") as f:
            hdr = f.readline().strip().split(",")
            n = sum(1 for _ in f)
        date_cols = [c for c in hdr if "date" in c.lower() or "pdufa" in c.lower()]
        print(f"  {name}: {n} rows")
        print(f"    date-ish columns: {date_cols[:12]}")

print("\n" + "=" * 92)
print("  D) armed_watchlist — conference lane we are not mining for DATES")
print("=" * 92)
aw = json.load(open(os.path.join(HERE, "Momentum Scanner", "armed_watchlist.json"),
                    encoding="utf-8", errors="replace"))
import collections
lanes = collections.Counter((v.get("lane") or "?") for v in aw["armed"].values())
print(f"  built {aw.get('built')} | sources {aw.get('sources')}")
print(f"  lanes: {dict(lanes)}")
conf = [(k, v) for k, v in aw["armed"].items() if (v.get("lane") or "") == "CONFERENCE"]
print(f"  CONFERENCE entries: {len(conf)}")
for k, v in conf[:6]:
    print(f"    {k:<7}{str(v.get('when')):<12}{str(v.get('why'))[:60]}")

print("\n" + "=" * 92)
print("  E) other files that carry dates")
print("=" * 92)
for pat in ("conference_trades_*.json", "readout_watchlist.txt", "catalyst_ctgov_cache.json",
            "iv_expansion_curve_results.json"):
    for p in glob.glob(os.path.join(HERE, pat)) + glob.glob(
            os.path.join(HERE, "Momentum Scanner", pat)):
        print(f"  {os.path.basename(p)}  {os.path.getsize(p)//1024}KB")
