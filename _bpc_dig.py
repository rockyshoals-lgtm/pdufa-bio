"""Two questions the headline comparison raised.

A) The 2026-10-23 cluster (16 names) and 10-24 (7). Real dates or another default? If the
   Conference column is populated, these are CONFERENCE PRESENTATION dates — genuinely hard,
   publicly announced, and exactly the class of date we want (the conference-signal study
   found 90.2% positive rate on these).
B) The PDUFA disagreements. A PDUFA is an FDA-ASSIGNED date; it is knowable and there is one
   right answer. Where BPC and we differ on a PDUFA, one of us is simply wrong, and that is
   the most damaging kind of error to publish.
"""
import csv, os, sys, collections, datetime
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.abspath(__file__))

ws = load_workbook(os.path.join(HERE, "bpc_data", "fda_2026-08-22.xlsx"),
                   read_only=True).worksheets[0]
hdr, rows = None, []
for r in ws.iter_rows(values_only=True):
    if hdr is None:
        hdr = [str(x or "").strip() for x in r]
        continue
    rows.append(dict(zip(hdr, r)))


def ds(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    return str(v or "")[:10]


print("=" * 96)
print("  A) THE 2026-10-23 / 10-24 CLUSTER — conference dates, or another default?")
print("=" * 96)
for target in ("2026-10-23", "2026-10-24", "2026-09-30", "2026-12-31"):
    sub = [d for d in rows if ds(d.get("Catalyst Date")) == target]
    conf = collections.Counter(str(d.get("Conference") or "(blank)") for d in sub)
    print(f"\n  {target}  n={len(sub)}   Conference column: {dict(list(conf.items())[:4])}")
    for d in sub[:5]:
        print(f"     {str(d.get('Ticker')):<7}{str(d.get('Stage'))[:12]:<13}"
              f"conf={str(d.get('Conference') or '-')[:22]:<24}"
              f"cat={str(d.get('Catalyst') or '')[:40]}")

print("\n" + "=" * 96)
print("  B) PDUFA DISAGREEMENTS — an FDA-assigned date has ONE right answer")
print("=" * 96)


def load(p):
    fp = os.path.join(HERE, p)
    return list(csv.DictReader(open(fp, encoding="utf-8-sig", errors="replace"))) \
        if os.path.exists(fp) else []


F, C, G = load("readout_forward.csv"), load("readout_calendar.csv"), load("ctgov_readouts.csv")
ours = collections.defaultdict(dict)
for r in F:
    if (r.get("window") or "").strip():
        ours[r["ticker"]]["EDGAR"] = r["window"].strip()
    if (r.get("window_alt") or "").strip():
        ours[r["ticker"]]["armed(pdufa.bio)"] = r["window_alt"].strip()
for r in G:
    if r.get("ticker") and (r.get("pcd") or "").strip():
        ours[r["ticker"]]["CTgov"] = r["pcd"].strip()

pd_rows = [d for d in rows if "PDUFA" in str(d.get("Stage") or "").upper()
           and ds(d.get("Catalyst Date")) >= "2026-08-22"]
print(f"  BPC forward PDUFA rows: {len(pd_rows)}\n")
print(f"  {'BPC date':<12}{'tk':<7}{'drug':<28} our sources")
for d in sorted(pd_rows, key=lambda x: ds(x.get("Catalyst Date")))[:26]:
    tk = str(d.get("Ticker") or "").upper()
    o = ours.get(tk)
    mine = ", ".join(f"{k}={v}" for k, v in o.items()) if o else "— WE HAVE NOTHING —"
    flag = ""
    if o and ds(d.get("Catalyst Date")) not in o.values():
        flag = "  <-- DIFFERS"
    print(f"  {ds(d.get('Catalyst Date')):<12}{tk:<7}{str(d.get('Drug') or '')[:26]:<28}{mine}{flag}")
