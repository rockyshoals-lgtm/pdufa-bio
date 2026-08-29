"""readout_gold_dates.py — the PUBLISHABLE date set. Precision-honest, source-attributed.

WHY THIS EXISTS (2026-08-22)
----------------------------
We nearly published 278 "hard dates" that were quarter buckets in disguise (31 tickers all
on 2026-12-31). BPC's export has the SAME disease, worse: 553 of its 735 forward rows sit on
New Year's Eve, 89% are placeholder-shaped. So "just use BPC's dates" would import 553 fake
dates. Neither source is trustworthy in bulk.

But inside BPC's file there are two classes of date that ARE hard, because they come from
outside the vendor's own guesswork:

  1. CONFERENCE-DATED rows. 15 of the 16 rows on 2026-10-23 carry ESMO26 in the Conference
     column -- a published congress agenda. Publicly announced, verifiable, and per our own
     conference study the highest-signal catalyst class we track (90.2% positive vs 76.7%).
  2. PDUFA rows. The FDA ASSIGNS these. There is exactly one right answer per application,
     so a PDUFA date is checkable rather than forecast.

Everything else stays a bucket and is rendered as a bucket.

OUTPUT (readout_gold_dates.csv), one row per ticker+event:
  date, precision (DAY|MONTH|QUARTER|HALF), confidence (GOLD|FIRM|SOFT), source, conflict
GOLD  = conference agenda or PDUFA, corroborated or single-sourced-but-external
FIRM  = a company stated the day itself in an SEC filing
SOFT  = bucket. Render as "Q4 2026" or "September 2026". NEVER as a day.

conflict != "" means two sources disagree on a checkable date -- surfaced, never silently
resolved. Those are the rows a human should look at before they go on the website.
"""
import csv, os, sys, collections, datetime, re

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.abspath(__file__))
TODAY = datetime.date.today()


def ds(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    return str(v or "")[:10]


def is_placeholder(d):
    """1st / 15th / month-end == the vendor bucket convention, not a real day."""
    try:
        dd = datetime.date.fromisoformat(d)
    except Exception:
        return True
    last = (datetime.date(dd.year + (dd.month == 12), dd.month % 12 + 1, 1)
            - datetime.timedelta(days=1)).day
    return dd.day in (1, 15, last)


def load_csv(p):
    fp = os.path.join(HERE, p)
    return list(csv.DictReader(open(fp, encoding="utf-8-sig", errors="replace"))) \
        if os.path.exists(fp) else []


# ---------------------------------------------------------------- BPC (the external anchors)
GOLD, rows_out = [], []
import glob
xs = sorted(glob.glob(os.path.join(HERE, "bpc_data", "fda_*.xlsx")))
if xs:
    from openpyxl import load_workbook
    ws = load_workbook(xs[-1], read_only=True).worksheets[0]
    hdr, brows = None, []
    for r in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(x or "").strip() for x in r]
            continue
        brows.append(dict(zip(hdr, r)))
    print(f"BPC: {os.path.basename(xs[-1])}, {len(brows)} rows")
    for d in brows:
        tk = str(d.get("Ticker") or "").strip().upper()
        date = ds(d.get("Catalyst Date"))
        if not tk or not re.match(r"^\d{4}-\d{2}-\d{2}$", date) or date < TODAY.isoformat():
            continue
        conf = str(d.get("Conference") or "").strip()
        stage = str(d.get("Stage") or "").strip()
        is_pdufa = "PDUFA" in stage.upper()
        if conf:
            GOLD.append({"ticker": tk, "date": date, "precision": "DAY",
                         "confidence": "GOLD", "source": f"BPC/conference:{conf[:38]}",
                         "event": stage, "drug": str(d.get("Drug") or "")[:40],
                         "note": str(d.get("Catalyst") or "")[:90]})
        elif is_pdufa and not is_placeholder(date):
            GOLD.append({"ticker": tk, "date": date, "precision": "DAY",
                         "confidence": "GOLD", "source": "BPC/PDUFA",
                         "event": stage, "drug": str(d.get("Drug") or "")[:40],
                         "note": str(d.get("Catalyst") or "")[:90]})
        elif is_pdufa:
            GOLD.append({"ticker": tk, "date": date, "precision": "MONTH",
                         "confidence": "SOFT", "source": "BPC/PDUFA-bucketed",
                         "event": stage, "drug": str(d.get("Drug") or "")[:40],
                         "note": str(d.get("Catalyst") or "")[:90]})

# ---------------------------------------------------------------- our conference miner
# 2026-08-29: conference_miner.py (EDGAR presenter announcements + conf_registry dates)
# replaces the hand-downloaded BPC export as the FRESH conference source. Same trust logic
# as the BPC conference column: the date comes from a published congress agenda, so an
# OBSERVED day is GOLD. A PROJECTED date (registry extrapolation, no agenda seen yet) is
# only FIRM at day precision and SOFT below — honest about what we have actually verified.
for r in load_csv("conference_presenters.csv"):
    tk, iso = r.get("ticker"), (r.get("catalyst_date") or "").strip()
    if not tk or not iso or iso[:10] < TODAY.isoformat():
        continue
    prec = (r.get("date_precision") or "").strip().lower()
    basis = (r.get("date_basis") or "").strip().lower()
    if prec == "day" and basis == "observed":
        conf_tier, out_prec = "GOLD", "DAY"
    elif prec == "day":
        conf_tier, out_prec = "FIRM", "DAY"
    else:
        conf_tier, out_prec = "SOFT", "MONTH"
    GOLD.append({"ticker": tk, "date": iso[:10] if out_prec == "DAY" else iso[:7],
                 "precision": out_prec, "confidence": conf_tier,
                 "source": f"EDGAR/conference:{r.get('conference', '')}",
                 "event": (r.get("pres_type") or "presentation"),
                 "drug": "", "note": (r.get("company") or "")[:60]
                 + (f" abst {r['abstract']}" if r.get("abstract") else "")})

# ---------------------------------------------------------------- ours
F = load_csv("readout_forward.csv")
G = load_csv("ctgov_readouts.csv")
for r in F:
    tk, w = r.get("ticker"), (r.get("window") or "").strip()
    if not tk or not w:
        continue
    p = (r.get("window_precision") or "").strip()
    if not p:                                   # pre-fix file: derive
        p = "DAY" if re.match(r"^[A-Z][a-z]+\.?\s+\d{1,2},?\s+20\d{2}$", w) else \
            ("QUARTER" if re.match(r"^Q[1-4]\s+20\d{2}$", w, re.I) else
             ("HALF" if re.match(r"^([12]H|MID)\s+20\d{2}$", w, re.I) else "OTHER"))
    if p == "DAY":
        try:
            iso = datetime.datetime.strptime(w.replace(",", ""), "%B %d %Y").date().isoformat()
        except Exception:
            iso = w
        GOLD.append({"ticker": tk, "date": iso, "precision": "DAY", "confidence": "FIRM",
                     "source": "EDGAR/company-stated", "event": "readout guidance",
                     "drug": "", "note": (r.get("phrases") or "")[:90]})
    else:
        GOLD.append({"ticker": tk, "date": w, "precision": p, "confidence": "SOFT",
                     "source": "EDGAR/guidance", "event": "readout guidance",
                     "drug": "", "note": (r.get("phrases") or "")[:90]})
for r in G:
    tk, pcd = r.get("ticker"), (r.get("pcd") or "").strip()
    if not tk or not pcd:
        continue
    prec = (r.get("pcd_precision") or "").strip()
    if not prec:
        prec = "MONTH" if is_placeholder(pcd) else "DAY"
    GOLD.append({"ticker": tk, "date": pcd if prec == "DAY" else pcd[:7],
                 "precision": prec,
                 "confidence": "SOFT",        # a data LOCK is not a PR date
                 "source": f"CTgov/{r.get('pcd_conf','')}", "event": "primary completion",
                 "drug": "", "note": (r.get("title") or "")[:90]})

# ---------------------------------------------------------------- conflicts
# CONFLICT, keyed on the EVENT not the ticker. First cut flagged 35 "conflicts" and most
# were not: MRK genuinely has four separate PDUFAs (WINREVAIR 9/21, WELIREG 10/04,
# ifinatamab 10/10, plus 10/27), and IONS has a cardiology talk on 8/28, another on 8/31,
# a PDUFA on 9/22 and company guidance for 10/26 -- four real events, zero disagreement.
# A conflict is only two sources dating the SAME DRUG differently. Rows without a drug
# name (our EDGAR/CT.gov rows) can only conflict against a BPC row for the same ticker
# when that ticker has exactly one BPC event, otherwise we cannot attribute them.
def _dkey(g):
    d = re.sub(r"[^a-z0-9]", "", (g.get("drug") or "").lower())[:14]
    return (g["ticker"], d)


by_ev = collections.defaultdict(list)
for g in GOLD:
    by_ev[_dkey(g)].append(g)
# attribute our undrugged rows to a ticker's ONLY event, if it has exactly one
tick_events = collections.defaultdict(set)
for (tk, d), gs in by_ev.items():
    if d:
        tick_events[tk].add(d)
for (tk, d), gs in list(by_ev.items()):
    if not d and len(tick_events.get(tk, ())) == 1:
        only = next(iter(tick_events[tk]))
        by_ev[(tk, only)].extend(gs)
        by_ev[(tk, d)] = []

for k, gs in by_ev.items():
    days = {g["date"] for g in gs if g["precision"] == "DAY"}
    if len(days) > 1:
        for g in gs:
            if g["precision"] == "DAY":
                g["conflict"] = "|".join(sorted(days - {g["date"]}))
for g in GOLD:
    g.setdefault("conflict", "")

GOLD.sort(key=lambda g: (g["date"], g["ticker"]))
dst = os.path.join(HERE, "readout_gold_dates.csv")
cols = ["date", "ticker", "precision", "confidence", "source", "event", "drug",
        "conflict", "note"]
try:
    f = open(dst, "w", newline="", encoding="utf-8")
except PermissionError:
    dst = dst.replace(".csv", "_new.csv")
    f = open(dst, "w", newline="", encoding="utf-8")
with f:
    w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
    w.writeheader()
    w.writerows(GOLD)

c = collections.Counter(g["confidence"] for g in GOLD)
p = collections.Counter(g["precision"] for g in GOLD)
print(f"\n{len(GOLD)} dated rows -> {os.path.basename(dst)}")
print(f"  confidence: {dict(c)}")
print(f"  precision : {dict(p)}")
nconf = sum(1 for g in GOLD if g["conflict"])
print(f"  conflicts flagged for human review: {nconf}")

print("\n" + "=" * 96)
print("  GOLD, next 60 days — hard enough to preload AND to publish")
print("=" * 96)
lim = (TODAY + datetime.timedelta(days=60)).isoformat()
for g in GOLD:
    if g["confidence"] == "GOLD" and g["date"] <= lim:
        print(f"  {g['date']}  {g['ticker']:<7}{g['event'][:16]:<18}{g['drug'][:26]:<28}"
              f"{g['source'][:40]}{'  ⚠CONFLICT ' + g['conflict'] if g['conflict'] else ''}")
