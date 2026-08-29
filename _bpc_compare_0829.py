"""Fresh BPC export (2026-08-29) vs our pipeline — the resilience audit.

Three questions, in order of how much they matter:
  1. WHAT MOVED in BPC's file since 8/22? Date changes on names we have preloaded are the
     most dangerous kind of staleness — we would be preloading against a dead date.
  2. WHAT does BPC have (dated, near-term) that OUR miners missed entirely — and is the miss
     a phrase gap, an alias gap, or a registry gap? Each gap class has a different fix.
  3. WHAT do WE have that BPC lacks? (Our edge — worth knowing, nothing to fix.)
"""
import csv, io, os, sys, re, glob, collections, datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
R = r"C:\Users\dcmoo\Documents\Python\9realms"
TODAY = datetime.date(2026, 8, 29)
from openpyxl import load_workbook


def bpc(path):
    ws = load_workbook(path, read_only=True).worksheets[0]
    hdr, rows = None, []
    for r in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(x or "").strip() for x in r]
            continue
        rows.append(dict(zip(hdr, r)))
    return rows


def ds(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    return str(v or "")[:10]


NEW = bpc(os.path.join(R, "bpc_data", "fda_2026-08-29.xlsx"))
OLD = bpc(os.path.join(R, "bpc_data", "fda_2026-08-22.xlsx"))
print(f"BPC 8/29: {len(NEW)} rows   8/22: {len(OLD)} rows")


def key(d):
    drug = re.sub(r"[^a-z0-9]", "", str(d.get("Drug") or "").lower())[:20]
    return (str(d.get("Ticker") or "").strip().upper(), drug)


old_by = {key(d): d for d in OLD}
new_by = {key(d): d for d in NEW}

# ---------------------------------------------------------------- 1. date drift 8/22 -> 8/29
print("\n" + "=" * 96)
print("  1) DATE DRIFT since 8/22 (same ticker+drug, different Catalyst Date) — future only")
print("=" * 96)
moved = []
for k, n in new_by.items():
    o = old_by.get(k)
    if not o:
        continue
    dn, do = ds(n.get("Catalyst Date")), ds(o.get("Catalyst Date"))
    if dn != do and dn >= TODAY.isoformat() and re.match(r"\d{4}-\d{2}", dn):
        moved.append((k[0], k[1][:18], do, dn, str(n.get("Stage") or "")[:16]))
moved.sort(key=lambda x: x[3])
for m in moved[:30]:
    print(f"    {m[0]:<7}{m[1]:<20}{m[2]}  ->  {m[3]}   {m[4]}")
print(f"  {len(moved)} future dates moved")

added = [n for k, n in new_by.items() if k not in old_by
         and ds(n.get("Catalyst Date")) >= TODAY.isoformat()]
gone = [o for k, o in old_by.items() if k not in new_by
        and ds(o.get("Catalyst Date")) >= TODAY.isoformat()]
print(f"  {len(added)} future rows ADDED since 8/22, {len(gone)} REMOVED")

# ---------------------------------------------------------------- 2. near-term BPC vs us
print("\n" + "=" * 96)
print("  2) BPC dated rows next 45d that OUR pipeline lacks (miss classification)")
print("=" * 96)


def load_csv(p):
    fp = os.path.join(R, p)
    return list(csv.DictReader(io.open(fp, encoding="utf-8-sig", errors="replace",
                                       newline=""))) if os.path.exists(fp) else []


ours = collections.defaultdict(set)          # ticker -> {source class}
for r in load_csv("readout_gold_dates.csv"):
    ours[r["ticker"]].add(r["source"].split("/")[0])
conf_ours = {r["ticker"] for r in load_csv("conference_presenters.csv")}

# what congress names appear in BPC that our alias table cannot resolve?
sys.path.insert(0, R)
import conference_presentations as CP
reg = CP.load_registry()

lim = (TODAY + datetime.timedelta(days=45)).isoformat()
missed, unresolved_conf = [], collections.Counter()
for n in NEW:
    tk = str(n.get("Ticker") or "").strip().upper()
    d = ds(n.get("Catalyst Date"))
    if not tk or not re.match(r"\d{4}-\d{2}-\d{2}", d) or not (TODAY.isoformat() <= d <= lim):
        continue
    conf = str(n.get("Conference") or "").strip()
    stage = str(n.get("Stage") or "")
    if conf:
        c = CP.detect_conference(conf)
        if not c:
            unresolved_conf[conf[:44]] += 1
        if tk not in conf_ours:
            missed.append((d, tk, "CONF", conf[:30], "alias" if not c else
                           ("no-filing-or-phrase" if c else "?")))
    elif "PDUFA" in stage.upper():
        if "BPC" not in ours.get(tk, set()) and "EDGAR" not in ours.get(tk, set()):
            missed.append((d, tk, "PDUFA", str(n.get("Drug") or "")[:30], "pdufa-source"))
missed.sort()
for m in missed[:35]:
    print(f"    {m[0]}  {m[1]:<7}{m[2]:<7}{m[3]:<32}{m[4]}")
print(f"  {len(missed)} near-term BPC rows not in our miner output")

print(f"\n  CONFERENCE NAMES our alias table cannot resolve (registry/alias gap -> fix these):")
for k, v in unresolved_conf.most_common(20):
    print(f"    {v:>2}x  {k}")

# ---------------------------------------------------------------- 3. ours that BPC lacks
print("\n" + "=" * 96)
print("  3) OUR events BPC's fresh file still lacks (the edge)")
print("=" * 96)
bpc_tks = {str(n.get("Ticker") or "").strip().upper() for n in NEW}
for r in load_csv("conference_presenters.csv"):
    if r["ticker"] not in bpc_tks:
        print(f"    {r['catalyst_date']}  {r['ticker']:<7}{r['conference']:<8}"
              f"{r['pres_type']:<13}{r['company'][:34]}  (not in BPC at all)")
    else:
        # in BPC but with a conference row?
        has_conf = any(str(n.get("Conference") or "").strip()
                       and str(n.get("Ticker") or "").strip().upper() == r["ticker"]
                       for n in NEW)
        if not has_conf:
            print(f"    {r['catalyst_date']}  {r['ticker']:<7}{r['conference']:<8}"
                  f"{r['pres_type']:<13}{r['company'][:34]}  (BPC has ticker, no conference row)")
