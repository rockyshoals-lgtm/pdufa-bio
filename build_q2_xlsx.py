"""Build companion Excel workbook with all scored Q2 catalysts + rotation blocks."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCORED = "/sessions/confident-serene-ptolemy/mnt/9realms/q2_scored_full.csv"
TIERED = "/sessions/confident-serene-ptolemy/mnt/9realms/q2_tiered_portfolio.csv"
OUT = "/sessions/confident-serene-ptolemy/mnt/9realms/Q2_2026_Scored_Universe.xlsx"

scored = pd.read_csv(SCORED)
tiered = pd.read_csv(TIERED)

wb = Workbook()
ws = wb.active
ws.title = "Summary"

# Styles
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2E5C8A")
zebra_fill = PatternFill("solid", fgColor="F2F4F7")
money_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", size=11, bold=True)
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# --- Summary sheet
ws["A1"] = "9 REALMS — Q2 2026 AGGRESSIVE STRATEGY"
ws["A1"].font = Font(name="Arial", size=16, bold=True, color="2E5C8A")
ws.merge_cells("A1:F1")

ws["A3"] = "Metric"; ws["B3"] = "Value"
for c in ["A3","B3"]:
    ws[c].font = header_font; ws[c].fill = header_fill; ws[c].alignment = center; ws[c].border = border

summary_rows = [
    ("Dated", "April 17, 2026"),
    ("Scope", "Q2 2026 (April 17 – June 30)"),
    ("Total H1 catalysts processed", 827),
    ("Q2 catalysts in window", 324),
    ("Aggressive-filter tiered roster size", len(tiered)),
    ("ALPHA tier (Gungnir)", int((tiered['tier']=="ALPHA").sum())),
    ("T1 tier (ODIN)", int((tiered['tier']=="T1").sum())),
    ("T2 tier (ODIN)", int((tiered['tier']=="T2").sum())),
    ("BETA tier (Gungnir)", int((tiered['tier']=="BETA").sum())),
    ("Nano/Micro-cap names", int(tiered['mcap_tier'].isin(['nano','micro']).sum())),
    ("Small-cap names", int((tiered['mcap_tier']=='small').sum())),
    ("Mid-cap names", int((tiered['mcap_tier']=='mid').sum())),
    ("If sized fully (naive sum)", f"{tiered['position_size_pct'].sum():.1f}%"),
    ("Target peak concurrent heat", "35–40%"),
    ("Target Q2 compounded return (base)", "60–120%"),
    ("Max drawdown kill-switch", "-12%"),
    ("Engines used", "ODIN v14 proxy, Gungnir v46 proxy, BIFROST v4 + v5.5, Conference, Smart Money, UOA, IIS"),
    ("Honest AUC recalibration", "ODIN ×0.96, Gungnir ×0.93"),
]
for i, (k, v) in enumerate(summary_rows, start=4):
    ws.cell(row=i, column=1, value=k).font = bold_font
    ws.cell(row=i, column=2, value=v).font = money_font
    for c in ["A","B"]:
        ws[f"{c}{i}"].border = border
        ws[f"{c}{i}"].alignment = left

ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 60

# --- Tiered Roster sheet
ws2 = wb.create_sheet("Tiered Roster")
display_cols = [
    "Ticker","Name","Drug","cat_class","catalyst_date","days_to_cat","mcap_tier","price",
    "ta","designations","conference","tier","inv_score_final","prob_positive_honest",
    "smart_money_flags","explosion_tier","iis_tier","bifrost_action",
    "entry_window","exit_window","position_size_pct","catalyst_text"
]
t = tiered[display_cols].copy()
for col_idx, col in enumerate(t.columns, 1):
    cell = ws2.cell(row=1, column=col_idx, value=col)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = border
for row_idx, row in enumerate(t.itertuples(index=False), 2):
    for col_idx, val in enumerate(row, 1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.font = money_font
        c.border = border
        c.alignment = left
        if row_idx % 2 == 0:
            c.fill = zebra_fill
# Widths
widths = [9, 32, 32, 12, 13, 10, 10, 10, 14, 20, 32, 9, 13, 13, 22, 14, 12, 14, 18, 24, 12, 60]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

# --- Full Scored Universe sheet
ws3 = wb.create_sheet("Full Scored (324)")
for col_idx, col in enumerate(scored.columns, 1):
    c = ws3.cell(row=1, column=col_idx, value=col)
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
for row_idx, row in enumerate(scored.itertuples(index=False), 2):
    for col_idx, val in enumerate(row, 1):
        c = ws3.cell(row=row_idx, column=col_idx, value=val)
        c.font = money_font
        c.border = border
        if row_idx % 2 == 0:
            c.fill = zebra_fill
for i in range(1, len(scored.columns)+1):
    ws3.column_dimensions[get_column_letter(i)].width = 16
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = ws3.dimensions

# --- Rotation Blocks sheet
ws4 = wb.create_sheet("Rotation Blocks")
rotation_rows = [
    ["BLOCK","DATES","DAYS","CATALYSTS","PEAK HEAT","KEY EXITS","KEY ENTRIES","COMMENTS"],
    ["Block 0 (Baseline)","Apr 17 open","1","4 held positions","95%","—","ALXO 55%, CMPX 40%, GRCE, WHWK, CRDF, CABA all held","Existing book — no new deployment yet"],
    ["Block 1 — AACR Week","Apr 17–22","5","18 ALPHA AACR plays","35%","WHWK Apr 21, CRDF Apr 18, CABA Apr 19","MOLN, FATE, TNXP, XLO, ZNTL, OLMA, AVBP, CNTX, PRLD, ACRV, FATE, COGT, ZLAB, HCM, AAPG + 3 nanos","21 ALPHA plays in 5 days — the fastest rotation. Staggered limit orders Apr 13–18."],
    ["Block 2 — PDUFA Double","Apr 23 – May 1","9","GRCE + AXSM + (skip ATYR)","14%","GRCE Apr 22 close","AXSM T-14 (Apr 16) options overlay","GRCE held from Block 0. AXSM mid-cap BTD = clean IV pattern."],
    ["Block 3 — May Microburst","May 2–May 31","30","ALXO ESMO + EDSA ATS + CING HEALEY + MNKD ATTD + CMPX","65% (dominated by ALXO+CMPX core)","ALXO May 6, CMPX May 27","EDSA T-14 May 6, MNKD T-14 May 16, CING T-21 May 10","Core positions exit, freeing 95% capital for Block 4"],
    ["Block 4 — June Cluster","Jun 1–Jun 30","30","ARVN, VRDN, UNCY, LNTH, DBVT, RNA, XNCR, OSTX, LRMR + CABA EULAR","38%","Everything Jun 27–30","CABA EULAR T-21 May 13, VRDN T-60 Apr 6, ARVN T-60 Apr 6","Deepest concentration block — 4 PDUFAs + 3 readouts in final week"],
]
for row_idx, row in enumerate(rotation_rows, 1):
    for col_idx, val in enumerate(row, 1):
        c = ws4.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
        else:
            c.font = money_font
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if row_idx % 2 == 1: c.fill = zebra_fill
widths4 = [22, 18, 6, 28, 12, 34, 42, 50]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
for r in range(2, len(rotation_rows)+1):
    ws4.row_dimensions[r].height = 50

# --- Current Holdings sheet
ws5 = wb.create_sheet("Current Holdings")
holdings = [
    ["Ticker","Weight","Catalyst","Date","Proxy Tier","Proxy Score","Action","Notes"],
    ["ALXO","55.0%","ESMO Breast belantamab","May 7","ALPHA","~100","HOLD → Exit T-1 May 6","Core position, biggest single bet"],
    ["CMPX","40.0%","ESMO GI pasritamig","May 28","ALPHA/BETA","~72","HOLD → Exit T-1 May 27","Second core, HER2+ GC binary"],
    ["Cash","5.0%","—","—","—","—","REDEPLOY Apr 17","Feed into Block 1 AACR entries"],
    ["GRCE","(added)","PDUFA GTX-104 orphan SAH","Apr 23","T2","89","HOLD → Exit Apr 22 close","Micro-cap PDUFA"],
    ["WHWK","(AACR)","AACR Oral × 3","Apr 17–22","ALPHA","~95","HOLD → Exit Apr 21 close","Multiple podium presentations"],
    ["CRDF","(AACR)","AACR Onvansertib combo","Apr 19","ALPHA","99","HOLD → Exit Apr 18 close","Combo with paclitaxel, novel data"],
    ["CABA","(AAN)","AAN RESET-MG + H1 SLE/SSc + EULAR Jun 3-6","Apr 20 / Jun 3","ALPHA","~85","HOLD → Exit Apr 19 for AAN, RELOAD T-21 for EULAR","100% MG-ADL response + BTD+ODD+RMAT"],
]
for row_idx, row in enumerate(holdings, 1):
    for col_idx, val in enumerate(row, 1):
        c = ws5.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
        else:
            c.font = money_font
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if row_idx % 2 == 1: c.fill = zebra_fill
widths5 = [10, 10, 34, 20, 14, 14, 40, 48]
for i, w in enumerate(widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.row_dimensions[1].height = 30

wb.save(OUT)
print(f"Wrote {OUT}")
