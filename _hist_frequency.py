"""HISTORICAL FREQUENCY + PHRASE MINING — historical_2026-07-18.xlsx (4,177 catalysts, 2025+).

David: "research how many of these happen a week... the historical file tells whether the trial
met its endpoint so you can compare... how often they happen, and more phrases to parse for."

TWO OUTPUTS, no fetching yet (that is the reaction step):
  1. FREQUENCY  — classify every historical catalyst, count phase readouts per week
  2. PHRASES    — mine the catalyst descriptions for readout language we don't yet parse

Honest framing up front: this file is BiopharmaCatalyst's HISTORICAL log. It is one vendor's
view, it has the date errors David already flagged, and "catalyst" here mixes readouts with
PDUFAs, approvals, conference talks and CRLs. So step 1 is separating the phase readouts from
everything else — the frequency number is only as good as that classification.
"""
import collections
import datetime as dt
import os
import re
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
U = (r"C:\Users\dcmoo\AppData\Roaming\Claude\local-agent-mode-sessions"
     r"\73ed6afa-1982-4aa5-beaa-ae356aeb0ed6\91666954-12a2-40a1-872a-dee734870139"
     r"\local_92dc8303-3ed0-4541-bb97-f41c446875d6\uploads")

wb = openpyxl.load_workbook(os.path.join(U, "historical_2026-07-18.xlsx"), read_only=True)
ws = wb.active
hdr = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
idx = {h: i for i, h in enumerate(hdr)}
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    rows.append(r)
wb.close()
print(f"{len(rows)} historical catalyst rows\n")


def g(r, k):
    i = idx.get(k)
    return r[i] if i is not None and i < len(r) else None


def as_date(x):
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    try:
        return dt.date(*map(int, str(x)[:10].split("-")))
    except Exception:
        return None


# ---- CLASSIFY -------------------------------------------------------------------------------
# The `Catalyst` text is free-form. Bucket it. A PHASE READOUT is the thing David trades:
# trial data reported/expected. NOT a PDUFA/approval (that is ODIN's job) and NOT a pure
# conference-scheduling note.
READOUT = re.compile(
    r"\bphase\s*[123]|topline|top-line|\bdata\b|readout|read-out|primary endpoint|"
    r"interim|cohort|met the|missed|did not meet|results?\b|analysis\b|proof of concept", re.I)
PDUFA = re.compile(r"\bPDUFA\b|\bNDA\b|\bBLA\b|\bsNDA\b|\b505\(b\)|complete response|"
                   r"\bCRL\b|approv|\bANDA\b|accepted for (review|filing)|priority review", re.I)
CONF_ONLY = re.compile(r"present|poster|oral|abstract|conference|symposium|late-break", re.I)


def bucket(cat, stage):
    c = (cat or "")
    s = (stage or "")
    if PDUFA.search(c) or "approv" in s.lower():
        return "PDUFA/approval"
    if re.search(r"phase\s*[123]", c, re.I) or re.search(r"PHASE", s, re.I):
        if READOUT.search(c):
            return "PHASE READOUT"
    if READOUT.search(c):
        return "PHASE READOUT"
    if CONF_ONLY.search(c):
        return "conference-only"
    return "other"


buck = collections.Counter()
readouts = []
for r in rows:
    d = as_date(g(r, "Catalyst Date"))
    cat = g(r, "Catalyst")
    stage = g(r, "Stage")
    b = bucket(cat, stage)
    buck[b] += 1
    if b == "PHASE READOUT" and d:
        readouts.append((d, g(r, "Ticker"), stage, cat, g(r, "Price At Catalyst Date")))

print("=" * 92)
print("  CATALYST TYPE BREAKDOWN (all 4,177)")
print("=" * 92)
for k, v in buck.most_common():
    print(f"  {v:>5}  {k}   ({v/len(rows)*100:.0f}%)")

# ---- FREQUENCY OF PHASE READOUTS ------------------------------------------------------------
readouts.sort()
dates = [d for d, *_ in readouts]
if dates:
    span_days = (max(dates) - min(dates)).days or 1
    weeks = span_days / 7
    print("\n" + "=" * 92)
    print("  PHASE-READOUT FREQUENCY")
    print("=" * 92)
    print(f"  {len(readouts)} phase readouts from {min(dates)} to {max(dates)} "
          f"({span_days} days = {weeks:.0f} weeks)")
    print(f"  -> {len(readouts)/weeks:.1f} phase readouts per week (ALL caps, ALL stages)")

    # by stage
    print("\n  by stage:")
    st = collections.Counter()
    for d, tk, stage, cat, px in readouts:
        s = (stage or "?").upper()
        s = "PHASE 3" if "3" in s else "PHASE 2" if "2" in s else "PHASE 1" if "1" in s else s
        st[s] += 1
    for s, n in st.most_common():
        print(f"    {s:<12} {n:>4}  ({n/weeks:.1f}/wk)")

    # by month — is it steady or lumpy?
    print("\n  by month (readouts):")
    mo = collections.Counter((d.year, d.month) for d, *_ in readouts)
    for (y, m), n in sorted(mo.items()):
        bar = "#" * int(n / max(mo.values()) * 40)
        print(f"    {y}-{m:02d}  {n:>4}  {bar}")

# ---- PHRASE MINING --------------------------------------------------------------------------
print("\n" + "=" * 92)
print("  PHRASE MINING — the real language in the catalyst column")
print("=" * 92)
# n-grams from readout descriptions, to find phrasing we don't parse
grams = collections.Counter()
for d, tk, stage, cat, px in readouts:
    words = re.findall(r"[a-z][a-z\-]+", (cat or "").lower())
    for n in (2, 3):
        for i in range(len(words) - n + 1):
            grams[" ".join(words[i:i + n])] += 1
print("  most common 2-3 word phrases in PHASE READOUT descriptions:")
for ph, n in grams.most_common(40):
    if n >= 8:
        print(f"    {n:>4}  {ph}")

# outcome language — does the text say met/missed?
print("\n  OUTCOME LANGUAGE present in the readout text:")
met = sum(1 for _, _, _, c, _ in readouts if re.search(r"\bmet\b|positive|success|achiev", c or "", re.I))
miss = sum(1 for _, _, _, c, _ in readouts if re.search(r"missed|did not meet|fail|negative", c or "", re.I))
print(f"    met/positive language : {met}")
print(f"    missed/fail language  : {miss}")
print(f"    neither (neutral desc): {len(readouts)-met-miss}")
print("  NOTE: most historical entries are logged NEUTRALLY (e.g. 'Phase 3 XYZ data') — the")
print("  outcome is not in the text. Reaction size (next step) is how we infer met vs missed.")
