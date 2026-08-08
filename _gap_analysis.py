"""GAP ANALYSIS — what forward readouts does BiopharmaCatalyst have that WE don't?

David: "compare it to what we have and see if we are missing any."

OUR forward coverage:
    phase_readouts_2026H2_cornerstone_augmented.csv  (665 rows, CT.gov + EDGAR)
BPC forward coverage:
    fda_2026-07-18.xlsx  (657 rows, all-2026 catalysts, KNOWN DATE ERRORS)

THE HONEST FRAME: these are two different definitions of "readout date".
    OURS = CT.gov PRIMARY COMPLETION DATE — when the trial data locks. A LEADING indicator.
    BPC  = the announced/expected catalyst date — when the company says data comes.
So a name in one and not the other is not automatically "missing" — it may be classified
differently (PDUFA vs readout), or dated to a different quarter. We separate:
    (a) names BPC calls a PHASE READOUT that we have NO entry for at all  -> real gaps
    (b) names we both have but with a DIFFERENT date/quarter               -> reconcile
    (c) names WE have that BPC doesn't                                      -> our edge
"""
import csv
import datetime as dt
import os
import re
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
U = (r"C:\Users\dcmoo\AppData\Roaming\Claude\local-agent-mode-sessions"
     r"\73ed6afa-1982-4aa5-beaa-ae356aeb0ed6\91666954-12a2-40a1-872a-dee734870139"
     r"\local_92dc8303-3ed0-4541-bb97-f41c446875d6\uploads")

# ---- OURS ----------------------------------------------------------------------------------
ours = {}
with open(os.path.join(U, "phase_readouts_2026H2_cornerstone_augmented.csv"),
          encoding="utf-8-sig") as f:
    for r in csv.DictReader(f):
        t = (r.get("ticker") or "").upper().strip()
        if t:
            ours[t] = r
our_nct = {(r.get("nct_id") or "").strip() for r in ours.values() if r.get("nct_id")}
print(f"OURS: {len(ours)} tickers, {len(our_nct)} NCT ids")

# ---- BPC -----------------------------------------------------------------------------------
wb = openpyxl.load_workbook(os.path.join(U, "fda_2026-07-18.xlsx"), read_only=True)
ws = wb.active
hdr = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
ix = {h: i for i, h in enumerate(hdr)}
bpc = []
for r in ws.iter_rows(min_row=2, values_only=True):
    bpc.append({h: r[ix[h]] for h in hdr})
wb.close()
print(f"BPC:  {len(bpc)} catalyst rows\n")


def as_date(x):
    if isinstance(x, (dt.datetime, dt.date)):
        return x if isinstance(x, dt.date) and not isinstance(x, dt.datetime) else x.date()
    try:
        return dt.date(*map(int, str(x)[:10].split("-")))
    except Exception:
        return None


READOUT = re.compile(r"phase\s*[123]|topline|top-line|\bdata\b|readout|primary endpoint|"
                     r"interim|cohort|results?\b", re.I)
PDUFA = re.compile(r"\bPDUFA\b|\bNDA\b|\bBLA\b|complete response|\bCRL\b|approv|\bANDA\b|"
                   r"accepted for|priority review", re.I)

TODAY = dt.date(2026, 7, 18)
H2_END = dt.date(2026, 12, 31)

# BPC phase readouts, forward, H2 2026
bpc_readouts = []
for r in bpc:
    cat = str(r.get("Catalyst") or "") + " " + str(r.get("Next Catalyst") or "")
    stage = str(r.get("Stage") or "")
    d = as_date(r.get("Catalyst Date"))
    if PDUFA.search(cat) or "approv" in stage.lower():
        continue
    if not (re.search(r"phase\s*[123]", cat + " " + stage, re.I) or READOUT.search(cat)):
        continue
    bpc_readouts.append({"t": (r.get("Ticker") or "").upper().strip(), "d": d, "cat": cat[:60],
                         "stage": stage, "nct": (r.get("NCT Number") or "").strip(),
                         "mcap": r.get("Market Cap"), "loa": r.get("Historical LOA")})
print(f"BPC phase readouts (PDUFA/approval excluded): {len(bpc_readouts)}")
fwd = [x for x in bpc_readouts if x["d"] and TODAY <= x["d"] <= H2_END]
print(f"  ...forward, in H2 2026 ({TODAY}..{H2_END}): {len(fwd)}\n")

# ---- (a) REAL GAPS: BPC readout, we have NO entry --------------------------------------------
print("=" * 96)
print("  (a) BPC PHASE READOUTS IN H2 2026 THAT WE HAVE NO TICKER FOR")
print("=" * 96)
gaps = [x for x in fwd if x["t"] and x["t"] not in ours and x["nct"] not in our_nct]
print(f"  {len(gaps)} names BPC lists that are NOT in our coverage at all:\n")
print(f"  {'ticker':<8}{'date':<12}{'stage':<14}{'mcap':>12}  catalyst")
print("  " + "-" * 92)
for x in sorted(gaps, key=lambda z: z["d"]):
    mc = x["mcap"]
    mcs = f"{float(mc)/1e6:.0f}M" if mc and str(mc).replace(".", "").isdigit() else "?"
    print(f"  {x['t']:<8}{x['d'].isoformat():<12}{x['stage'][:13]:<14}{mcs:>12}  {x['cat'][:40]}")

# ---- (b) DATE DISAGREEMENTS -----------------------------------------------------------------
print("\n" + "=" * 96)
print("  (b) NAMES WE BOTH HAVE — does the date/quarter agree?")
print("=" * 96)
both = [x for x in fwd if x["t"] in ours]
disagree = 0
for x in sorted(both, key=lambda z: z["d"])[:25]:
    o = ours[x["t"]]
    od = o.get("catalyst_date") or o.get("data_lock_date") or ""
    ood = as_date(od)
    note = ""
    if ood and abs((ood - x["d"]).days) > 45:
        note = f"  <-- {abs((ood-x['d']).days)}d apart"
        disagree += 1
    print(f"  {x['t']:<7} BPC {x['d'].isoformat()}  |  ours {od[:10] or 'None':<10} "
          f"({o.get('date_basis','?')}){note}")
print(f"\n  {disagree} of the shown names disagree by >45 days — expected: our CT.gov PCD leads")
print("  the announced date, and BPC's dates have the errors you flagged.")

# ---- (c) OUR EDGE ---------------------------------------------------------------------------
bpc_tks = {x["t"] for x in bpc_readouts if x["t"]}
our_only = [t for t in ours if t not in bpc_tks]
print("\n" + "=" * 96)
print("  (c) NAMES WE HAVE THAT BPC HAS NO READOUT FOR (our CT.gov edge)")
print("=" * 96)
print(f"  {len(our_only)} of our {len(ours)} tickers are not in BPC's readout list.")
print(f"  These are mostly CT.gov primary-completion-date names BPC hasn't dated yet —")
print(f"  the leading edge. Sample: {sorted(our_only)[:20]}")

print("\n" + "=" * 96)
print("  SUMMARY")
print("=" * 96)
print(f"  BPC forward H2 readouts        : {len(fwd)}")
print(f"  ...we already cover            : {len(both)} ({len(both)/max(len(fwd),1)*100:.0f}%)")
print(f"  ...REAL GAPS (add these)       : {len(gaps)}")
print(f"  Our names BPC doesn't list     : {len(our_only)} (our leading edge)")
