# -*- coding: utf-8 -*-
"""Build the H2 phase-readout workbook from the miner CSV.

WHY THIS EXISTS
`phase_readouts_2026H2_CLEAN.csv` (2026-07-12, 2,323 rows) was 67% junk: 1,250 RECRUITING +
302 NOT_YET_RECRUITING. A trial still enrolling patients has not read out and will not read out
on its stated date. The miner was fixed on 2026-07-12/13 -- every run since has zero enrolling
rows -- but the workbook was never rebuilt, so the contamination lived on downstream.

THE ONE THING THIS WORKBOOK MUST DO
Never let a CT.gov ESTIMATE be mistaken for a date the COMPANY actually stated.

Those are different animals and they are the whole reason a readout calendar goes wrong:
  * ctgov_pcd          -- an ESTIMATED primary completion date a sponsor typed into a registry
                          and is under no obligation to hit. It slips. It is a proxy, not a date.
  * company_guidance   -- what the company TOLD THE MARKET in an SEC filing ("topline expected
                          in Q4"). Precision is usually a quarter/half, but the commitment is real.
  * conference_schedule / pr -- a named session or a scheduling PR. The only EXACT DAY there is,
                          and vanishingly rare: measured lead time is a median of T-3 (n=4 across
                          205 tickers), so do not expect this column to be populated.

So every row carries date_basis + date_precision + confidence, and the date cell is COLOUR-CODED
by provenance. An amber date is a guess with a number on it. Sorting by date alone is a mistake
this workbook is built to make hard.

`redistribute=False` rows (the FMP newswire leg) are highlighted red and listed on the README.
FMP's redistribution terms are still unread (P2-6), so nothing flagged False may reach the site,
the API, or the sitemap. This workbook is INTERNAL.

Usage:  python build_readout_xlsx.py [--csv phase_readouts_2026H2.csv] [--out phase_readouts_2026H2.xlsx]
"""
import argparse, datetime as dt, os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
FONT = 'Arial'

# provenance -> (fill, meaning). The point of the whole file.
BASIS_FILL = {
    'company_guidance':    ('C6EFCE', 'Company told the market (SEC filing)'),
    'conference_schedule': ('BDD7EE', 'Named conference session — exact day'),
    'pr':                  ('BDD7EE', 'Company scheduling PR — exact day'),
    'fmp_press':           ('BDD7EE', 'Company scheduling PR — exact day'),
    'ctgov_pcd':           ('FFE699', 'CT.gov ESTIMATE — sponsor-typed, slips, proxy only'),
}
PENDING_FILL = 'D9D9D9'
RED = 'FFC7CE'

COLS = ['ticker', 'catalyst_date', 'date_precision', 'date_basis', 'confidence', 'imminence',
        'days_to_readout', 'readout_stage', 'data_lock_date', 'redistribute', 'phase', 'status',
        'drug', 'indication', 'trial', 'nct_id', 'sponsor', 'enrollment', 'milestone',
        'conference', 'pres_type', 'source', 'source_url', 'note']


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', fgColor='1F3864')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def build(csv, out):
    d = pd.read_csv(csv, low_memory=False)

    # ---- integrity gates. Fail loudly rather than ship a contaminated workbook. ----
    if 'status' in d.columns:
        bad = d[d.status.astype(str).str.upper().str.contains('RECRUIT', na=False)
                & ~d.status.astype(str).str.upper().str.contains('NOT_RECRUIT', na=False)]
        if len(bad):
            raise SystemExit(f'ABORT: {len(bad)} still-enrolling rows in {csv} — this is exactly '
                             f'the bug the workbook exists to keep out. Re-run the miner without '
                             f'--include-enrolling.')
    if 'imminence' in d.columns and (d.imminence == 'DISTANT').any():
        n = int((d.imminence == 'DISTANT').sum())
        raise SystemExit(f'ABORT: {n} DISTANT rows — those merely stopped recruiting and are not '
                         f'upcoming readouts. Re-run with --imminent-days 90.')

    # PAST — stale company guidance whose stated window has already closed ("topline expected Q2
    # 2026", read in July). --imminent-days does NOT remove these: its horizon test is
    # `days_to_readout <= N`, and a PAST row's days_to_readout is NEGATIVE, so it passes. The
    # miner's own tier legend calls PAST "window closed (should be 0)". Dropped here, loudly --
    # a date that has already come and gone is not an upcoming readout, and leaving it in an
    # upcoming-readouts file is how a stale row gets traded.
    n_past = 0
    if 'imminence' in d.columns:
        n_past = int((d.imminence == 'PAST').sum())
        if n_past:
            d = d[d.imminence != 'PAST']
            print(f'  dropped {n_past} PAST rows (stated window already closed — not upcoming)')

    # ---- MONTH-END PLACEHOLDER DEMOTION -------------------------------------------------
    # date_precision='day' does NOT mean "we know the day". It means "the CT.gov field happened
    # to contain a day". Measured on the 2026-07-16 run (n=397 ctgov_pcd day-precision rows):
    #     31st: 79   30th: 62   1st: 26   -> 42.1% on the 1st/30th/31st
    # A uniform distribution would put ~10% there. So ~4x enrichment: these are sponsors typing
    # "2026-12-31" to mean "sometime in Q4", not a trial reading out on New Year's Eve. Worse,
    # 12-31 and 06-30 are exactly the dates a naive reader treats as hard.
    #
    # Every day-precision row in that run was ctgov_pcd; ZERO company_guidance rows had a day.
    # So a "day" here is never a company commitment -- it is always a registry estimate, and
    # 4 times in 10 it is a placeholder. Demote those to month and say so, rather than let a
    # placeholder be colour-coded and sorted as if it were a date.
    if {'date_basis', 'date_precision', 'catalyst_date'} <= set(d.columns):
        _dt = pd.to_datetime(d.catalyst_date, errors='coerce')
        placeholder = (
            (d.date_basis == 'ctgov_pcd') & (d.date_precision == 'day')
            & _dt.notna() & _dt.dt.day.isin([1, 30, 31])
            # month-end only: the 30th/31st of a 30/31-day month, or the 1st. Feb 28 is a real
            # month end but also a plausible real date, so it is left alone.
        )
        n_ph = int(placeholder.sum())
        if n_ph:
            d.loc[placeholder, 'date_precision'] = 'month'
            d.loc[placeholder, 'note'] = (
                d.loc[placeholder, 'note'].fillna('').astype(str)
                + ' | CT.gov day-precision demoted to month: lands on the 1st/30th/31st, the '
                  'signature of a month-end placeholder rather than a stated day.').str.strip(' |')
            print(f'  demoted {n_ph} CT.gov month-end placeholder dates from day -> month precision')

    d = d.reindex(columns=[c for c in COLS if c in d.columns])
    order = {'OVERDUE': 0, 'IMMINENT': 1, 'NEAR': 2, 'SCHEDULED': 3, 'UNDATED': 4, 'PAST': 5}
    d['_o'] = d.imminence.map(order).fillna(9)
    d = d.sort_values(['_o', 'days_to_readout'], na_position='last').drop(columns='_o')

    wb = Workbook()

    # ---------------- README ----------------
    rd = wb.active
    rd.title = 'README'
    stamp = dt.date.today().isoformat()
    n_pr = int((d.redistribute == False).sum()) if 'redistribute' in d.columns else 0  # noqa: E712
    lines = [
        ('H2 2026 PHASE READOUTS — upcoming only', True),
        (f'Built {stamp} from {os.path.basename(csv)} · {len(d)} rows', False),
        ('', False),
        ('WHAT IS IN HERE', True),
        ('Only trials that can actually read out soon. Two states qualify:', False),
        ('  enrollment_closed  — ACTIVE_NOT_RECRUITING, last patient in, primary completion ahead.', False),
        ('  completed_pending  — primary completion PASSED, no results posted. Topline overdue.', False),
        ('                       These have NO catalyst_date: the passed date is the DATA LOCK,', False),
        ('                       not the readout. data_lock_date holds it. Watchlist, not calendar.', False),
        ('', False),
        ('WHAT IS DELIBERATELY NOT IN HERE', True),
        ('Trials still enrolling (RECRUITING / NOT_YET_RECRUITING). A trial still taking patients', False),
        ('has not read out and will not read out on its stated date. It is a pipeline entry, not a', False),
        ('catalyst. The previous workbook (phase_readouts_2026H2_CLEAN.csv, 2026-07-12) was 67%', False),
        ('these — 1,552 of 2,323 rows. The miner was fixed 2026-07-12/13; this is the first rebuild.', False),
        ('DISTANT trials (>180d) are also excluded: closed enrollment alone is not imminence.', False),
        ('', False),
        ('WHY \'day\' PRECISION IS NOT A DAY', True),
        ('date_precision="day" means the CT.gov field held a day, NOT that the day is known.', False),
        ('On the 2026-07-16 run 42.1% of CT.gov day-precision dates fell on the 1st/30th/31st', False),
        ('(vs ~10% expected) — the signature of a month-end placeholder. Those are demoted to', False),
        ('"month" at build. No company_guidance row has EVER carried day precision, so a day', False),
        ('here is always a registry estimate, never a commitment.', False),
        ('', False),
        ('READ THE DATE COLOUR BEFORE YOU READ THE DATE', True),
        ('Not all dates are the same kind of fact:', False),
        ('  GREEN  company_guidance   — the company TOLD the market (SEC filing). Real commitment,', False),
        ('                              usually quarter/half precision.', False),
        ('  BLUE   conference_schedule / pr — a named session or scheduling PR. The only EXACT DAY.', False),
        ('                              Rare by nature: measured lead time is a MEDIAN OF T-3', False),
        ('                              (n=4 across 205 tickers). Do not expect many.', False),
        ('  AMBER  ctgov_pcd          — an ESTIMATE the sponsor typed into a registry and is under', False),
        ('                              no obligation to hit. It slips. A proxy, not a date.', False),
        ('  GREY   pending            — no date at all (topline overdue). Use data_lock_date.', False),
        ('Sorting by catalyst_date alone silently mixes commitments with guesses. Use date_basis.', False),
        ('', False),
        (f'{n_past} PAST row(s) were dropped at build: company guidance whose stated window has', False),
        ('already closed. --imminent-days does not catch these (a negative days_to_readout still', False),
        ('passes a "<= 90" test), so they are removed here. A date already gone is not upcoming.', False),
        ('', False),
        ('REDISTRIBUTION — INTERNAL FILE', True),
        (f'{n_pr} row(s) are redistribute=False (FMP newswire leg). FMP redistribution terms are', False),
        ('still unread (P2-6). Nothing flagged False may reach the site, the API, or the sitemap.', False),
        ('Those rows are filled RED in the Readouts tab. This workbook is internal.', False),
        ('', False),
        ('Informational and educational only — not investment advice. Verify every date against', False),
        ('primary sources (company IR / SEC) before acting on it.', False),
    ]
    for i, (t, bold) in enumerate(lines, start=1):
        c = rd.cell(row=i, column=1, value=t)
        c.font = Font(name=FONT, bold=bold, size=11 if bold else 10)
    rd.column_dimensions['A'].width = 104

    # ---------------- Readouts ----------------
    ws = wb.create_sheet('Readouts')
    ws.append(list(d.columns))
    for _, r in d.iterrows():
        ws.append([None if pd.isna(v) else v for v in r.tolist()])
    style_header(ws, len(d.columns))

    ci = {c: i + 1 for i, c in enumerate(d.columns)}
    thin = Side(style='thin', color='D9D9D9')
    for row in range(2, len(d) + 2):
        for c in range(1, len(d.columns) + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical='top')
        basis = ws.cell(row=row, column=ci['date_basis']).value
        prec = ws.cell(row=row, column=ci['date_precision']).value
        dcell = ws.cell(row=row, column=ci['catalyst_date'])
        if prec == 'pending' or not dcell.value:
            dcell.fill = PatternFill('solid', fgColor=PENDING_FILL)
        elif basis in BASIS_FILL:
            dcell.fill = PatternFill('solid', fgColor=BASIS_FILL[basis][0])
        dcell.font = Font(name=FONT, size=10, bold=True)
        if 'redistribute' in ci and ws.cell(row=row, column=ci['redistribute']).value in (False, 'False'):
            ws.cell(row=row, column=ci['redistribute']).fill = PatternFill('solid', fgColor=RED)

    widths = {'ticker': 8, 'catalyst_date': 13, 'date_precision': 10, 'date_basis': 18,
              'confidence': 10, 'imminence': 11, 'days_to_readout': 9, 'readout_stage': 17,
              'data_lock_date': 13, 'redistribute': 11, 'phase': 12, 'status': 21, 'drug': 30,
              'indication': 34, 'trial': 16, 'nct_id': 13, 'sponsor': 28, 'enrollment': 10,
              'milestone': 16, 'conference': 12, 'pres_type': 10, 'source': 17,
              'source_url': 40, 'note': 60}
    for c, w in widths.items():
        if c in ci:
            ws.column_dimensions[get_column_letter(ci[c])].width = w

    # ---------------- Summary (formulas, not baked numbers) ----------------
    sm = wb.create_sheet('Summary')
    last = len(d) + 1
    imm_col = get_column_letter(ci['imminence'])
    bas_col = get_column_letter(ci['date_basis'])
    rng_i = f"Readouts!${imm_col}$2:${imm_col}${last}"
    rng_b = f"Readouts!${bas_col}$2:${bas_col}${last}"

    sm['A1'] = 'H2 2026 upcoming readouts — composition'
    sm['A1'].font = Font(name=FONT, bold=True, size=12)
    sm['A3'] = 'By imminence'; sm['A3'].font = Font(name=FONT, bold=True)
    sm['A4'] = 'OVERDUE — data locked, topline pending'
    sm['A5'] = 'IMMINENT — <=45d'
    sm['A6'] = 'NEAR — 46-90d'
    for i, key in enumerate(['OVERDUE', 'IMMINENT', 'NEAR'], start=4):
        sm[f'B{i}'] = f'=COUNTIF({rng_i},"{key}")'
    sm['A7'] = 'TOTAL'; sm['A7'].font = Font(name=FONT, bold=True)
    sm['B7'] = '=SUM(B4:B6)'; sm['B7'].font = Font(name=FONT, bold=True)

    sm['A9'] = 'By date provenance'; sm['A9'].font = Font(name=FONT, bold=True)
    rows = [('company_guidance', 'Company told the market (SEC)'),
            ('conference_schedule', 'Conference session — exact day'),
            ('fmp_press', 'Scheduling PR — exact day'),
            ('ctgov_pcd', 'CT.gov ESTIMATE — slips, proxy only')]
    for i, (k, lab) in enumerate(rows, start=10):
        sm[f'A{i}'] = lab
        sm[f'B{i}'] = f'=COUNTIF({rng_b},"{k}")'
        sm[f'C{i}'] = k
    sm['A14'] = 'TOTAL'; sm['A14'].font = Font(name=FONT, bold=True)
    sm['B14'] = '=SUM(B10:B13)'; sm['B14'].font = Font(name=FONT, bold=True)
    sm['A16'] = 'Company-stated share of dated rows'
    sm['B16'] = '=IFERROR((B10+B11+B12)/B14,0)'
    sm['B16'].number_format = '0.0%'
    sm['A17'] = ('Low is expected, not a defect: scheduling PRs give a median of T-3 warning '
                 '(n=4 across 205 tickers).')
    for r in range(1, 18):
        for c in ('A', 'B', 'C'):
            if sm[f'{c}{r}'].font.name != FONT:
                sm[f'{c}{r}'].font = Font(name=FONT, size=10)
    sm.column_dimensions['A'].width = 46
    sm.column_dimensions['B'].width = 12
    sm.column_dimensions['C'].width = 20

    wb.save(out)
    print(f'wrote {out}  ({len(d)} rows)')
    print('  imminence :', d.imminence.value_counts().to_dict())
    print('  date_basis:', d.date_basis.value_counts(dropna=False).to_dict())
    print(f'  redistribute=False rows: {n_pr} (internal only \u2014 must not reach site/API/sitemap)')
    return out


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--csv', default=os.path.join(HERE, 'phase_readouts_2026H2.csv'))
    ap.add_argument('--out', default=os.path.join(HERE, 'phase_readouts_2026H2.xlsx'))
    a = ap.parse_args()
    build(a.csv, a.out)
